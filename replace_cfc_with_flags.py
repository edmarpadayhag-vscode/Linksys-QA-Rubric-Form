
"""
Replace §6 Critical Failure Conditions with auto-populated Coaching Flags.
When a BI that was linked to a former CFC is rated Partial or Missed,
the corresponding flag lights up automatically. Multiple flags are listed.

Rows 131-144 (old §6) are reused in place — no row shifts needed.
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
import re

FILE  = r"d:\Visual Studio Code\QA Rubric\QA Scoring Form Template.xlsx"
SHEET = "Scoring Form"

# ── Positions confirmed from survey ──────────────────────────────
CFC_START = 131   # old §6 header row
CFC_END   = 144   # old "CRITICAL FAILURE FLAGGED?" row

# BI → coaching flag mapping
# (flag_row, description, BI_cell, coaching_focus)
FLAGS = [
    (133, "Technical accuracy failure",
     "C50",
     "Review for incorrect guidance, product misrepresentation, or unsupported root-cause conclusions"),
    (134, "No follow-up after disconnected call",
     "C76",
     "Mandatory callback attempt required; document outcome in case notes"),
    (135, "Commitments not fulfilled / clear path forward missing",
     "C73",
     "Review what was promised; confirm the customer was left with actionable next steps"),
    (136, "Customer abandoned without a path forward",
     "C38",
     "Review whether the customer had any defined next step at close; address in coaching session"),
    (137, "Escalated without completing reasonable troubleshooting",
     "C84",
     "Review L1 steps completed before escalation; identify gaps and add to troubleshooting playbook"),
    (138, "Escalation criteria not met / complaint or legal-risk not escalated",
     "C85",
     "Review whether a valid escalation trigger existed; coach on recognising escalation criteria"),
    (139, "Professionalism concern",
     "C103",
     "Mandatory review; assess tone, conduct, and whether a policy violation occurred"),
    (140, "Ticket status inaccurate / premature closure",
     "C119",
     "Review ticket state at close; verify status accurately reflected the case outcome"),
]
FLAG_ROWS  = [f[0] for f in FLAGS]
NOTE_ROW   = 141
SUMM_ROW   = 142
LIST_ROW   = 143
# Row 144 left as an empty spacer

# ── Style helpers ─────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def fnt(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic)

def aln(wrap=True, v="top", h="left"):
    return Alignment(wrap_text=wrap, vertical=v, horizontal=h)

_s = Side(style="thin", color="BFBFBF")
def bdr():
    return Border(left=_s, right=_s, top=_s, bottom=_s)

# ── Load ──────────────────────────────────────────────────────────
print("Loading...")
wb = load_workbook(FILE)
ws = wb[SHEET]

# ── 1. Unmerge & clear old §6 rows ───────────────────────────────
print(f"Clearing rows {CFC_START}–{CFC_END}...")
for m in [m for m in list(ws.merged_cells.ranges)
          if CFC_START <= m.min_row <= CFC_END]:
    ws.unmerge_cells(str(m))
for r in range(CFC_START, CFC_END + 1):
    for c in range(1, 8):
        ws.cell(r, c).value = None
        # Reset fill on B–F for clean slate
        ws.cell(r, c).fill = fill("FFFFFF")

# ── 2. Remove C133:C143 from the Yes/No DV ────────────────────────
print("Updating DVs...")
updated_dvs = []
for dv in ws.data_validations.dataValidation:
    sq = str(dv.sqref)
    if "C133" in sq and "C143" in sq:
        # Remove the C133:C143 range; keep C129 and C152:C162 and E9
        new_sq = sq.replace("C133:C143", "").strip()
        new_sq = re.sub(r'\s+', ' ', new_sq).strip()
        dv.sqref = new_sq
        print(f"  DV updated: removed C133:C143 → '{new_sq}'")
    updated_dvs.append(dv)
ws.data_validations.dataValidation = updated_dvs

# ── 3. Write new Coaching Flags section ───────────────────────────
print("Writing Coaching Flags section...")

# Row 131 – Section header
ws.merge_cells(start_row=131, start_column=2, end_row=131, end_column=6)
c = ws["B131"]
c.value = ("COACHING FLAGS   ·   Auto-populated from Partial and Missed ratings   ·   "
           "Each flagged item requires a coaching or review action")
c.font  = fnt(bold=True, color="FFFFFF", size=10)
c.fill  = fill("2E75B6")
c.alignment = aln(h="center", v="center")
c.border = bdr()
ws.row_dimensions[131].height = 22

# Row 132 – Column headers
for col, txt, span in [(2,"Flag Condition",1),(3,"Rating",1),(4,"Coaching Focus / Required Action",3)]:
    c = ws.cell(132, col)
    c.value = txt
    c.font  = fnt(bold=True, color="FFFFFF", size=9.5)
    c.fill  = fill("4472C4")
    c.alignment = aln(h="center", v="center")
    c.border = bdr()
if True:
    ws.merge_cells(start_row=132, start_column=4, end_row=132, end_column=6)
ws.row_dimensions[132].height = 18

# Rows 133–140 – Flag rows
for (row, desc, bi_cell, coaching) in FLAGS:
    # B – description
    c = ws.cell(row, 2)
    c.value     = desc
    c.font      = fnt(size=9.5)
    c.fill      = fill("FFF2CC")
    c.alignment = aln(v="center")
    c.border    = bdr()

    # C – auto-formula: returns the BI rating if Partial/Missed, else ""
    c = ws.cell(row, 3)
    c.value     = f'=IF(OR({bi_cell}="Partial",{bi_cell}="Missed"),{bi_cell},"")'
    c.font      = fnt(bold=True, size=9.5)
    c.fill      = fill("FFF2CC")
    c.alignment = aln(h="center", v="center")
    c.border    = bdr()

    # D:F – coaching focus (merged)
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
    c = ws.cell(row, 4)
    c.value     = coaching
    c.font      = fnt(size=9.5, italic=True)
    c.fill      = fill("FFF2CC")
    c.alignment = aln(v="center")
    c.border    = bdr()

    ws.row_dimensions[row].height = 28

# Row 141 – Survey / CSAT note (no matching BI)
ws.merge_cells(start_row=NOTE_ROW, start_column=2, end_row=NOTE_ROW, end_column=6)
c = ws.cell(NOTE_ROW, 2)
c.value     = ("⚠  Survey / CSAT manipulation has no matching behavioral indicator and "
               "cannot be auto-detected — flag manually in §8 Coaching & Feedback if suspected.")
c.font      = fnt(size=9, italic=True, color="7B4F00")
c.fill      = fill("FFFBEC")
c.alignment = aln(h="left", v="center")
c.border    = bdr()
ws.row_dimensions[NOTE_ROW].height = 22

# Row 142 – Count summary
fr0, fr1 = FLAG_ROWS[0], FLAG_ROWS[-1]
summ_formula = (
    f'=IF(COUNTIF(C{fr0}:C{fr1},"Partial")+COUNTIF(C{fr0}:C{fr1},"Missed")=0,'
    f'"✓  No coaching flags triggered — all key indicators met",'
    f'COUNTIF(C{fr0}:C{fr1},"Partial")+COUNTIF(C{fr0}:C{fr1},"Missed")'
    f'&" coaching flag(s) triggered — review and coaching action required")'
)
ws.merge_cells(start_row=SUMM_ROW, start_column=2, end_row=SUMM_ROW, end_column=6)
c = ws.cell(SUMM_ROW, 2)
c.value     = summ_formula
c.font      = fnt(bold=True, size=10)
c.fill      = fill("D9E1F2")
c.alignment = aln(h="center", v="center")
c.border    = bdr()
ws.row_dimensions[SUMM_ROW].height = 22

# Row 143 – Triggered flags list (TEXTJOIN)
parts = [f'IF(C{r}<>"",'
         f'"⚑ "&B{r}&" ["&C{r}&"]","")'
         for r in FLAG_ROWS]
list_formula = (
    f'=IF(COUNTIF(C{fr0}:C{fr1},"Partial")+COUNTIF(C{fr0}:C{fr1},"Missed")=0,'
    f'"All key indicators met — no coaching action required",'
    f'TEXTJOIN(CHAR(10),TRUE,{",".join(parts)}))'
)
ws.merge_cells(start_row=LIST_ROW, start_column=2, end_row=LIST_ROW, end_column=6)
c = ws.cell(LIST_ROW, 2)
c.value     = list_formula
c.font      = fnt(size=9.5)
c.fill      = fill("FFFFFF")
c.alignment = aln(h="left", v="top")
c.border    = bdr()
ws.row_dimensions[LIST_ROW].height = 65

# Row 144 – spacer
ws.merge_cells(start_row=144, start_column=2, end_row=144, end_column=6)
ws.cell(144, 2).fill = fill("F8F8F8")
ws.row_dimensions[144].height = 6

# ── 4. Conditional formatting on C133:C140 ────────────────────────
print("Adding conditional formatting...")
cf_range = f"C{fr0}:C{fr1}"
ws.conditional_formatting.add(
    cf_range,
    CellIsRule(operator="equal", formula=['"Partial"'],
               fill=fill("FFC000"),
               font=Font(bold=True, color="7B4F00", size=9.5)))
ws.conditional_formatting.add(
    cf_range,
    CellIsRule(operator="equal", formula=['"Missed"'],
               fill=fill("C00000"),
               font=Font(bold=True, color="FFFFFF", size=9.5)))

# ── 5. Remove CFC COUNTIF from E30, F30, C163 ────────────────────
print("Patching weighted total and auto-zero formulas...")
cfc_countif = f',COUNTIF(C{fr0}:C{fr1},"Yes")>0'
# Also matches the old raw range reference pattern
cfc_old     = r',\s*COUNTIF\(C133:C143,"Yes"\)>0'

def strip_cfc(formula):
    f = re.sub(cfc_old, '', formula, flags=re.IGNORECASE)
    return f

cells_to_patch = [(30,5),(30,6)]
# Find auto-zero trigger
for r in range(155, 170):
    v = ws.cell(r, 3).value
    if v and 'Quality Score forced' in str(v):
        cells_to_patch.append((r, 3))
        print(f"  Auto-zero trigger at C{r}")
        break

for (row, col) in cells_to_patch:
    v = ws.cell(row, col).value
    if v and 'COUNTIF(C133' in str(v):
        ws.cell(row, col).value = strip_cfc(v)
        print(f"  Patched {chr(64+col)}{row}")

# ── 6. Update §7 Auto-Zero description (remove "§6" reference) ────
for r in range(144, 150):
    v = ws.cell(r, 2).value
    if v and '§6' in str(v):
        ws.cell(r, 2).value = str(v).replace(
            ', or any §6 critical-failure condition', '')
        print(f"  Removed §6 ref from row {r}")
        break

print("\nSaving...")
wb.save(FILE)
print("Done.")
