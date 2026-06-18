
"""
Option B — replace free 0-5 inputs with descriptive dropdowns for
Journey Outcome (E56:E59) and Continuity of Support (F56:F59).

Steps:
1. Lists sheet  — add Journey Outcome (col H) and Continuity (col I) options
2. Case Tracker — replace whole-number DVs with list DVs referencing Lists sheet
3. Case Tracker — update G56:G59 to extract numeric score from selected text
4. Case Tracker — update E55/F55 column headers
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy

FILE = r"d:\Visual Studio Code\QA Rubric\QA Scoring Form Template.xlsx"
wb   = load_workbook(FILE)

# ══════════════════════════════════════════════════════════════════
# 1. LISTS SHEET — add dropdown options
# ══════════════════════════════════════════════════════════════════
print("=== Lists sheet: adding Journey Outcome and Continuity options ===")
ws_l = wb["Lists"]

JO_OPTIONS = [
    "5 — Fully resolved and confirmed",
    "4 — Strong path: RMA / escalation / limitation explained",
    "3 — Partial — next step exists, issue still open",
    "2 — Minimal progress despite multiple touches",
    "1 — No meaningful progress",
    "0 — Worsened or contradictory guidance",
]

CONT_OPTIONS = [
    "5 — Full context preserved, no repeated questions",
    "4 — Minor gap at one handoff, no customer impact",
    "3 — Customer re-explained parts of the issue",
    "2 — Significant context loss across touches",
    "1 — Near-zero continuity, agents restarted fresh",
    "0 — Contradictory guidance, complete breakdown",
]

# Column H — Journey Outcome
ws_l.cell(1, 8).value = "Journey Outcome"
ws_l.cell(1, 8).font  = Font(bold=True, size=10)
for i, opt in enumerate(JO_OPTIONS, start=2):
    ws_l.cell(i, 8).value = opt
ws_l.column_dimensions["H"].width = 52
print(f"  H1:H{1+len(JO_OPTIONS)} — Journey Outcome options written")

# Column I — Continuity of Support
ws_l.cell(1, 9).value = "Continuity of Support"
ws_l.cell(1, 9).font  = Font(bold=True, size=10)
for i, opt in enumerate(CONT_OPTIONS, start=2):
    ws_l.cell(i, 9).value = opt
ws_l.column_dimensions["I"].width = 52
print(f"  I1:I{1+len(CONT_OPTIONS)} — Continuity options written")

# ══════════════════════════════════════════════════════════════════
# 2. CASE TRACKER — update DVs
# ══════════════════════════════════════════════════════════════════
print("\n=== Case Tracker: updating DVs ===")
ws_ct = wb["Case Tracker"]

# Remove the existing E and F DVs (whole-number + tooltip DVs added in Option A)
before = len(ws_ct.data_validations.dataValidation)
ws_ct.data_validations.dataValidation = [
    dv for dv in ws_ct.data_validations.dataValidation
    if not any(x in str(dv.sqref) for x in ["E56", "F56"])
]
after = len(ws_ct.data_validations.dataValidation)
print(f"  Removed {before - after} old DVs for E56/F56 range")

# Journey Outcome — list DV referencing Lists!$H$2:$H$7
dv_jo = DataValidation(
    type="list",
    formula1="Lists!$H$2:$H$7",
    allow_blank=True,
    showInputMessage=True,
    promptTitle="Journey Outcome",
    prompt="Select the option that best describes the overall resolution achieved across all touches.",
    showErrorMessage=True,
    errorTitle="Invalid selection",
    error="Please choose from the dropdown list.",
)
dv_jo.sqref = "E56:E59"
ws_ct.add_data_validation(dv_jo)
print("  Added Journey Outcome list DV → E56:E59")

# Continuity of Support — list DV referencing Lists!$I$2:$I$7
dv_cont = DataValidation(
    type="list",
    formula1="Lists!$I$2:$I$7",
    allow_blank=True,
    showInputMessage=True,
    promptTitle="Continuity of Support",
    prompt="Select the option that best describes how well context and ownership were maintained across handoffs.",
    showErrorMessage=True,
    errorTitle="Invalid selection",
    error="Please choose from the dropdown list.",
)
dv_cont.sqref = "F56:F59"
ws_ct.add_data_validation(dv_cont)
print("  Added Continuity list DV → F56:F59")

# ══════════════════════════════════════════════════════════════════
# 3. CASE TRACKER — update Team Score formula (G56:G59)
# ══════════════════════════════════════════════════════════════════
# E and F now hold text like "5 — Fully resolved..."
# Extract numeric score with: VALUE(LEFT(E56,1))
# Wrap in IFERROR in case cell is blank or unexpected value.

print("\n=== Case Tracker: updating Team Score formulas ===")

def team_score_formula(r):
    jo   = f'IFERROR(VALUE(LEFT(E{r},1)),0)'   # Journey Outcome score
    cont = f'IFERROR(VALUE(LEFT(F{r},1)),0)'   # Continuity score
    has_l1 = f'COUNTIFS($B$5:$B$49,A{r},$Q$5:$Q$49,"Level 1")>0'
    has_l2 = f'COUNTIFS($B$5:$B$49,A{r},$Q$5:$Q$49,"Level 2")>0'
    l1_avg = f'IFERROR(AVERAGEIFS($L$5:$L$49,$B$5:$B$49,A{r},$Q$5:$Q$49,"Level 1"),0)'
    l2_avg = f'IFERROR(AVERAGEIFS($L$5:$L$49,$B$5:$B$49,A{r},$Q$5:$Q$49,"Level 2"),0)'
    return (
        f'=IF(OR(A{r}="",D{r}="",E{r}="",F{r}=""),"",IFERROR('
        f'IF(AND({has_l1},{has_l2}),'
        f'0.2*{l1_avg}+0.3*{l2_avg},'
        f'0.5*D{r})'
        f'+0.3*({jo}/5)+0.2*({cont}/5),""))'
    )

for r in (56, 57, 58, 59):
    ws_ct.cell(r, 7).value = team_score_formula(r)
    print(f"  G{r}: updated")

# ══════════════════════════════════════════════════════════════════
# 4. CASE TRACKER — update column headers E55 and F55
# ══════════════════════════════════════════════════════════════════
print("\n=== Case Tracker: updating column headers ===")
old_e = ws_ct.cell(55, 5).value
old_f = ws_ct.cell(55, 6).value
ws_ct.cell(55, 5).value = "Journey Outcome"
ws_ct.cell(55, 6).value = "Continuity of Support"
print(f"  E55: '{old_e}' → 'Journey Outcome'")
print(f"  F55: '{old_f}' → 'Continuity of Support'")

# ── Save ───────────────────────────────────────────────────────────
print("\nSaving...")
wb.save(FILE)
print("Done.")
