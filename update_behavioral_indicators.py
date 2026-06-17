
import openpyxl
from openpyxl import load_workbook
import re

FILE  = r"d:\Visual Studio Code\QA Rubric\QA Scoring Form Template.xlsx"
SHEET = "Scoring Form"

# (row, col, new_text | None=clear)
# None = merged away; row cleared and removed from DV
CHANGES = [
    # ── §1 Resolution / Progress (rows 39-43 → 3 indicators) ──────
    (39, 2, "Resolved the issue or achieved the best possible outcome "
            "(RMA, valid escalation, limitation explained, education)"),
    (40, 2, "Completed reasonable troubleshooting before reaching a conclusion"),
    (41, 2, "Used the appropriate resolution path for the situation and warranty status, "
            "moving the customer closer to a solution with clear next steps."),
    (42, 2, None),   # merged into row 41
    (43, 2, None),   # merged into row 41

    # ── §2 Technical Accuracy (rows 52-58 → 4 indicators, reordered) ──
    # New order: merged(52+53), root-cause-moved-up(→53), merged(54+55), merged(56+57)
    (52, 2, "Correctly identified the symptoms and asked relevant diagnostic questions"),
    (53, 2, "Identified the root cause (or likely cause) when possible"),   # moved up
    (54, 2, "Followed a logical troubleshooting process, and available tools were used, "
            "efficiently and appropriately."),
    (55, 2, None),   # merged into row 54
    (56, 2, "Gave technically accurate information; no unsupported conclusions"),
    (57, 2, None),   # merged into row 56
    (58, 2, None),   # content moved to row 53; this row cleared

    # ── §3 Communication & Call Control (rows 67-74) ──────────────
    # Item 2 NOT implemented: rows 69 and 71 unchanged
    # (67)  unchanged: "Expectations were set clearly from the start"
    (68, 2, "Agent maintained call control and troubleshooting flow"),
    # (69)  UNCHANGED (item 2 skipped)
    (70, 2, "Actively listened and showed comprehension throughout the interaction/s"),
    # (71)  UNCHANGED (item 2 skipped)
    (72, 2, "Managed time well, with clear transitions between steps"),
    (73, 2, "Handled difficult moments when needed to maintain call control"),
    (74, 2, "Adjusted communication for the customer's accessibility needs "
            "(hearing, vision, language barrier, age, cognitive)"),

    # ── §4 Customer Ownership (rows 83-90 → 5 indicators) ────────
    # Item 3 NOT implemented: rows 87 and 88 unchanged
    (83, 2, "Took accountability for the customer's outcome and avoided unnecessary transfers"),
    (84, 2, None),   # merged into row 83
    (85, 2, "Commitments made were followed through, and a clear path forward was "
            "created for the customer."),
    (86, 2, None),   # merged into row 85
    # (87)  UNCHANGED (item 3 skipped): "Next actions were documented"
    # (88)  UNCHANGED (item 3 skipped): "Realistic expectations were set"
    (89, 2, "Completed required disconnect follow-up and met callback commitments "
            "within the agreed window"),
    (90, 2, None),   # merged into row 89

    # ── §5 Escalation Judgment (rows 99-103 → 4 indicators) ──────
    # Item 1 fix: "for a valid reason" removed from row 99 (kept as row 100)
    (99,  2, "Completed reasonable troubleshooting before escalating"),
    (100, 2, "Identified a valid reason to escalate"),
    (101, 2, "Provided complete escalation details using the correct escalation path"),
    (102, 2, None),  # merged into row 101
    (103, 2, "Clearly explained why escalation was necessary and what would happen next."),

    # ── §6 Customer Experience (rows 119-124 → 5 indicators) ─────
    (119, 2, "Demonstrated empathy including for repeat-contact fatigue"),
    (120, 2, "Stayed professional throughout and maintained patience, even under pressure"),
    (121, 2, "Agent matched the customer's tone, pace, level, and energy"),
    (122, 2, None),  # merged into row 120
    (123, 2, "Minimized customer effort"),
    (124, 2, "Maintained customer understanding and engagement"),

    # ── §7 Documentation & Notes (rows 133-139 → 6 indicators) ──
    (133, 2, "Recorded customer issue summary"),
    (134, 2, "Documented troubleshooting steps"),
    (135, 2, "Noted key findings"),
    (136, 2, "Captured agent reasoning - escalation justification, offline "
             "recommendations provided by colleagues and other support POCs "
             "especially if no private notes applicable to step/action."),
    (137, 2, None),  # merged into row 136
    (138, 2, "Documented next steps and follow-ups"),
    (139, 2, "Updated ticket status accurately"),
]

# Rows whose C-column cells are cleared and must be removed from the indicator DV
CLEARED_ROWS = {42, 43, 55, 57, 58, 84, 86, 90, 102, 122, 137}

# ── Load ──────────────────────────────────────────────────────────
print("Loading...")
wb = load_workbook(FILE)
ws = wb[SHEET]

# ── Apply text changes ────────────────────────────────────────────
print("Updating indicators...")
for row, col, text in CHANGES:
    ws.cell(row, col).value = text
    status = "CLEAR" if text is None else text[:60]
    print(f"  Row {row:3d}: {status}")

# ── Update behavioral indicator DV (remove cleared rows) ─────────
print("\nUpdating DV sqref...")
for dv in ws.data_validations.dataValidation:
    sqref_str = str(dv.sqref)
    # Identify the behavioral indicators DV by a known cell it contains
    if 'C39' in sqref_str:
        cells = re.findall(r'[A-Z]+\d+', sqref_str)
        kept = [c for c in cells
                if not (re.search(r'\d+', c) and
                        int(re.search(r'\d+', c).group()) in CLEARED_ROWS)]
        new_sqref = ' '.join(kept)
        dv.sqref = new_sqref
        print(f"  Old count: {len(cells)} cells")
        print(f"  New count: {len(kept)} cells")
        print(f"  Removed:   {sorted(CLEARED_ROWS)}")
        break

# ── Save ──────────────────────────────────────────────────────────
print("\nSaving...")
wb.save(FILE)
print("Done.")
