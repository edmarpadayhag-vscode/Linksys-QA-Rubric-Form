
from openpyxl import load_workbook

FILE  = r"d:\Visual Studio Code\QA Rubric\QA Scoring Form Template.xlsx"
wb = load_workbook(FILE)
ws = wb["Case Tracker"]

for r in (56, 57, 58, 59):
    ws.cell(r, 7).number_format = "0.00%"
    print(f"G{r}: number_format set to 0.00%")

wb.save(FILE)
print("Done.")
