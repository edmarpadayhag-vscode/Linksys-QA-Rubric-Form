
"""Remove Team Score Weight label (D5) and formula (E5:F5) from Agent Level row."""
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side
from copy import copy

FILE  = r"d:\Visual Studio Code\QA Rubric\QA Scoring Form Template.xlsx"
SHEET = "Scoring Form"

wb = load_workbook(FILE)
ws = wb[SHEET]

# Unmerge E5:F5
for m in list(ws.merged_cells.ranges):
    if m.min_row == 5 and m.min_col == 5:
        ws.unmerge_cells(str(m))
        print(f"  Unmerged {m}")

# Clear D5, E5, F5
for col in (4, 5, 6):
    c = ws.cell(5, col)
    c.value = None
    # Copy fill from C5 to keep row visually consistent
    c.fill  = copy(ws.cell(5, 3).fill)
    c.border = copy(ws.cell(5, 3).border)
    c.alignment = copy(ws.cell(5, 3).alignment)

print("Cleared D5, E5, F5")

wb.save(FILE)
print("Done.")
