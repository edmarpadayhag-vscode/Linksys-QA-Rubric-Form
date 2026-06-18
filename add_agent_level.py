
"""
Insert a new row 5 for "Agent Level" after the Agent Name row (row 4).
Old rows 5–188 shift to 6–189.

New row layout (matches existing Case Info rows):
  B5 : "Agent Level"          — label
  C5 : dropdown Level 1 / Level 2   — data entry
  D5 : "Team Score Weight"   — label
  E5:F5 merged : formula showing the weight

Team Score Weight logic:
  Level 1  →  1.0×  (standard weight in team aggregations)
  Level 2  →  1.5×  (senior support weight in team aggregations)
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from copy import copy
import re

FILE  = r"d:\Visual Studio Code\QA Rubric\QA Scoring Form Template.xlsx"
SHEET = "Scoring Form"

INSERT_AT = 5          # new row 5; old rows >= 5 shift by +1
OLD_MAX   = 188        # current max content row (inspect reported 188)
NEW_MAX   = OLD_MAX + 1

# ── Row mapping ────────────────────────────────────────────────────
def map_row(r):
    """Old row → new row. Rows < INSERT_AT stay; rows >= INSERT_AT shift +1."""
    if r < INSERT_AT:
        return r
    return r + 1

# ── Formula translator ─────────────────────────────────────────────
def translate_formula(formula):
    def sub(m):
        col, row = m.group(1), int(m.group(2))
        return f"{col}{map_row(row)}"
    return re.sub(r'([A-Z]+)(\d+)', sub, formula)

# ── sqref translator ───────────────────────────────────────────────
def remap_sqref(s):
    def sub(m):
        col, row = m.group(1), int(m.group(2))
        return f"{col}{map_row(row)}"
    return re.sub(r'([A-Z]+)(\d+)', sub, s)

# ── Style helpers ──────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def fnt(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic)

def aln(wrap=True, v="center", h="left"):
    return Alignment(wrap_text=wrap, vertical=v, horizontal=h)

_s = Side(style="thin", color="BFBFBF")
def bdr():
    return Border(left=_s, right=_s, top=_s, bottom=_s)

# ── Load ───────────────────────────────────────────────────────────
print("Loading...")
wb = load_workbook(FILE)
ws = wb[SHEET]

# ── Snapshot all rows >= INSERT_AT ────────────────────────────────
print(f"Snapshotting rows {INSERT_AT}–{OLD_MAX}...")
snap = {}
for row in ws.iter_rows(min_row=INSERT_AT, max_row=OLD_MAX):
    for cell in row:
        if cell.value is not None or cell.has_style:
            snap[(cell.row, cell.column)] = {
                'value': cell.value,
                'has_style': cell.has_style,
                **({k: copy(getattr(cell, k))
                    for k in ('font','fill','border','alignment',
                              'protection','number_format')}
                   if cell.has_style else {})
            }

row_heights = {rd.index: rd.height for rd in ws.row_dimensions.values() if rd.height}
merge_data  = [(m.min_row, m.min_col, m.max_row, m.max_col)
               for m in ws.merged_cells.ranges]
old_dvs     = list(ws.data_validations.dataValidation)

# Snapshot conditional formatting: extract (sqref_str, rules) pairs
# Keys are ConditionalFormatting objects with a .sqref attribute (MultiCellRange)
old_cf = [(str(k.sqref) if hasattr(k, 'sqref') else str(k), list(v))
          for k, v in ws.conditional_formatting._cf_rules.items()]

# ── Unmerge everything; clear rows INSERT_AT onward ───────────────
print("Clearing shifted rows...")
for m in list(ws.merged_cells.ranges):
    ws.unmerge_cells(str(m))
for row in ws.iter_rows(min_row=INSERT_AT, max_row=NEW_MAX + 1):
    for cell in row:
        cell.value = None
ws.data_validations.dataValidation = []

# ── Re-apply rows that don't move (rows 1..INSERT_AT-1) ──────────
# Merges for those rows are rebuilt below alongside the rest.

# ── Write shifted content ─────────────────────────────────────────
print("Writing shifted content...")
for (old_r, col), data in sorted(snap.items()):
    nr = map_row(old_r)
    dst = ws.cell(nr, col)
    val = data['value']
    if isinstance(val, str) and val.startswith('='):
        val = translate_formula(val)
    dst.value = val
    if data.get('has_style'):
        for attr in ('font','fill','border','alignment','protection','number_format'):
            if attr in data:
                setattr(dst, attr, data[attr])

# ── Rebuild merges ────────────────────────────────────────────────
print("Rebuilding merges...")
for (min_r, min_c, max_r, max_c) in merge_data:
    nm = map_row(min_r)
    nx = map_row(max_r)
    try:
        ws.merge_cells(start_row=nm, start_column=min_c,
                       end_row=nx,   end_column=max_c)
    except Exception as e:
        print(f"  merge warn {min_r}-{max_r}: {e}")

# ── Rebuild row heights ───────────────────────────────────────────
for old_r, h in row_heights.items():
    nr = map_row(old_r) if old_r >= INSERT_AT else old_r
    ws.row_dimensions[nr].height = h

# ── Write new Agent Level row (row 5) ────────────────────────────
print("Writing Agent Level row...")

# Copy style reference from row 4 (same Case Info row type)
ref_row = 4

# B5 – left label  (match B4 style)
c = ws.cell(5, 2)
c.value     = "Agent Level"
c.font      = copy(ws.cell(ref_row, 2).font)
c.fill      = copy(ws.cell(ref_row, 2).fill)
c.alignment = copy(ws.cell(ref_row, 2).alignment)
c.border    = bdr()

# C5 – left value  (data entry; match C4 style)
c = ws.cell(5, 3)
c.value     = None
c.font      = copy(ws.cell(ref_row, 3).font)
c.fill      = copy(ws.cell(ref_row, 3).fill)
c.alignment = copy(ws.cell(ref_row, 3).alignment)
c.border    = bdr()

# D5 – right label  (match D4 style)
c = ws.cell(5, 4)
c.value     = "Team Score Weight"
c.font      = copy(ws.cell(ref_row, 4).font)
c.fill      = copy(ws.cell(ref_row, 4).fill)
c.alignment = copy(ws.cell(ref_row, 4).alignment)
c.border    = bdr()

# E5:F5 merged – formula  (match E4:F4 style)
ws.merge_cells(start_row=5, start_column=5, end_row=5, end_column=6)
c = ws.cell(5, 5)
c.value     = ('=IF(C5="Level 2","1.5×  —  senior support (L2 weight applies to team aggregations)",'
               'IF(C5="Level 1","1.0×  —  standard support (L1 weight applies to team aggregations)",""))')
c.font      = fnt(bold=False, size=9.5, italic=True, color="2E4057")
c.fill      = copy(ws.cell(ref_row, 5).fill)
c.alignment = aln(h="left", v="center")
c.border    = bdr()

ws.row_dimensions[5].height = ws.row_dimensions[ref_row].height or 18

# ── Rebuild DVs with shifted sqrefs ──────────────────────────────
print("Rebuilding DVs...")
for dv in old_dvs:
    new_sqref = remap_sqref(str(dv.sqref))
    if not new_sqref.strip():
        continue
    ndv = DataValidation(
        type=dv.type, operator=dv.operator,
        formula1=dv.formula1, formula2=dv.formula2,
        allow_blank=dv.allow_blank,
        showInputMessage=dv.showInputMessage,
        showErrorMessage=dv.showErrorMessage,
        errorTitle=dv.errorTitle, error=dv.error,
        promptTitle=dv.promptTitle, prompt=dv.prompt,
    )
    ndv.sqref = new_sqref
    ws.add_data_validation(ndv)
    if new_sqref != str(dv.sqref):
        print(f"  {str(dv.sqref)[:55]!r} → {new_sqref[:55]!r}")

# ── Add new DV for C5 (Agent Level dropdown) ─────────────────────
print("Adding Agent Level DV...")
dv_level = DataValidation(
    type="list",
    formula1='"Level 1,Level 2"',
    allow_blank=True,
    showInputMessage=True,
    promptTitle="Agent Level",
    prompt="Select Level 1 for front-line support agents or Level 2 for senior/escalation agents.",
    showErrorMessage=True,
    errorTitle="Invalid selection",
    error="Please choose Level 1 or Level 2.",
)
dv_level.sqref = "C5"
ws.add_data_validation(dv_level)

# ── Rebuild conditional formatting with shifted sqrefs ────────────
print("Rebuilding conditional formatting...")
ws.conditional_formatting._cf_rules = {}
for cf_range_str, rules in old_cf:
    new_range_str = remap_sqref(cf_range_str)
    for rule in rules:
        ws.conditional_formatting.add(new_range_str, rule)
    print(f"  CF {cf_range_str!r} → {new_range_str!r}")

# ── Ghost cell cleanup ────────────────────────────────────────────
print("Cleaning ghost cells...")
to_del = [(r, c) for (r, c) in list(ws._cells.keys()) if r > NEW_MAX]
for key in to_del:
    del ws._cells[key]
print(f"  Removed {len(to_del)} ghost cells (max content row = {NEW_MAX})")

# ── Verify key cells after shift ─────────────────────────────────
print("\n=== Verification (key formula cells) ===")
check_cells = [
    (22, 4),   # D22 — Score Summary header row (was 22, stays 22 since < INSERT_AT=5... wait no)
    (24, 4),   # D24 — Resolution score (was D23, now D24)
    (31, 5),   # E31 — Weighted total (was E30, now E31)
    (32, 6),   # F32...
]
# Actually let's just print rows 21-35 relevant cells
for r in range(21, 36):
    for c in [2, 3, 4, 5, 6]:
        v = ws.cell(r, c).value
        if v is not None:
            print(f"  {chr(64+c)}{r}: {str(v)[:80]}")

print("\nSaving...")
wb.save(FILE)
print(f"Done. New max row ≈ {NEW_MAX}")
