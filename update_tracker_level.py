
"""
Case Tracker: add Level column and update Team Score formula.

Changes:
1. Q4  — "Level" header (copy style from P4)
2. Q5:Q49 — dropdown: Level 1 / Level 2
3. G56:G59 — replace 0.5*D with 0.2*L1_avg + 0.3*L2_avg
4. Row 54 — update description text
5. D55  — rename to "Avg Touch % (ref)" to make clear D is reference only
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy

FILE  = r"d:\Visual Studio Code\QA Rubric\QA Scoring Form Template.xlsx"
SHEET = "Case Tracker"

wb = load_workbook(FILE)
ws = wb[SHEET]

# ── Helpers ────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

_s = Side(style="thin", color="BFBFBF")
def bdr():
    return Border(left=_s, right=_s, top=_s, bottom=_s)

# ── 1. Column Q header (row 4) ────────────────────────────────────
print("Adding Level header Q4...")
ref = ws.cell(4, 16)          # P4 — copy style from here
c   = ws.cell(4, 17)          # Q4
c.value     = "Level"
c.font      = copy(ref.font)
c.fill      = copy(ref.fill)
c.alignment = copy(ref.alignment)
c.border    = copy(ref.border)
ws.column_dimensions["Q"].width = 14

# ── 2. Q5:Q49 — data entry cells ──────────────────────────────────
print("Styling Q5:Q49...")
ref_data = ws.cell(5, 16)     # P5 — copy data row style from here
for r in range(5, 50):
    c = ws.cell(r, 17)
    c.font      = copy(ref_data.font)
    c.fill      = copy(ref_data.fill)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = bdr()
    ws.row_dimensions[r].height = ws.row_dimensions[r].height or 18

# ── 3. DV for Q5:Q49 ──────────────────────────────────────────────
print("Adding Level DV...")
dv = DataValidation(
    type="list",
    formula1='"Level 1,Level 2"',
    allow_blank=True,
    showInputMessage=True,
    promptTitle="Agent Level",
    prompt="Level 1 = front-line support.  Level 2 = senior / escalation support.",
    showErrorMessage=True,
    errorTitle="Invalid",
    error="Please select Level 1 or Level 2.",
)
dv.sqref = "Q5:Q49"
ws.add_data_validation(dv)

# ── 4. Update Team Score formulas G56:G59 ────────────────────────
print("Updating Team Score formulas...")

# New formula template (row-parameterised)
# Old: 0.5*D{r}+0.3*(E{r}/5)+0.2*(F{r}/5)
# New: 0.2*L1_avg + 0.3*L2_avg + 0.3*(E/5) + 0.2*(F/5)
#      where L1_avg = AVERAGEIFS(L5:L49, B5:B49, A{r}, Q5:Q49, "Level 1")
#      and   L2_avg = AVERAGEIFS(L5:L49, B5:B49, A{r}, Q5:Q49, "Level 2")
def team_score_formula(r):
    return (
        f'=IF(OR(A{r}="",D{r}="",E{r}="",F{r}=""),"",IFERROR('
        f'0.2*IFERROR(AVERAGEIFS($L$5:$L$49,$B$5:$B$49,A{r},$Q$5:$Q$49,"Level 1"),0)'
        f'+0.3*IFERROR(AVERAGEIFS($L$5:$L$49,$B$5:$B$49,A{r},$Q$5:$Q$49,"Level 2"),0)'
        f'+0.3*(E{r}/5)+0.2*(F{r}/5),""))'
    )

for r in (56, 57, 58, 59):
    old = ws.cell(r, 7).value
    new = team_score_formula(r)
    ws.cell(r, 7).value = new
    print(f"  G{r}: updated")
    print(f"    was: {str(old)[:80]}")
    print(f"    now: {new[:80]}")

# ── 5. Update row 54 description text ────────────────────────────
print("\nUpdating row 54 description...")
old_desc = ws.cell(54, 1).value
new_desc = (
    "Team Score  =  20% × Level 1 avg touch  +  30% × Level 2 avg touch  "
    "+  30% × (Journey Outcome ÷ 5)  +  20% × (Continuity of Support ÷ 5)   "
    "│   Level 2 touches carry a higher weight in the team aggregate."
)
ws.cell(54, 1).value = new_desc
print(f"  was: {str(old_desc)[:90]}")
print(f"  now: {new_desc[:90]}")

# ── 6. Rename D55 to clarify it is a reference column ─────────────
print("\nUpdating D55 header...")
ws.cell(55, 4).value = "Avg Touch % (ref)"

# ── 7. Save ───────────────────────────────────────────────────────
print("\nSaving...")
wb.save(FILE)
print("Done.")
