
import openpyxl
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy
import re

FILE = r"d:\Visual Studio Code\QA Rubric\QA Scoring Form Template.xlsx"
SHEET = "Scoring Form"

# Rows 1-107       → delta 0    (unchanged)
# Row 108          → overwritten by Escalation Follow-Up block (rows 108-114)
# Rows 109-201     → delta +7   (CE + Doc shift; §4-§8 rows 188-201 also shift)
# NEW rows 209-212 → Next Session Focus (inserted between coaching rec and Q3)
# Rows 202-208     → delta +11  (Q3 through Final Assessment)
def get_delta(row):
    if row <= 107:   return 0
    if 109 <= row <= 201: return 7
    if 202 <= row <= 208: return 11
    return None

def map_row(row):
    if row <= 3: return row
    d = get_delta(row)
    return row + d if d is not None else None

# Score Summary cells (rows 24-33, no shift) need their cross-block refs updated
# because CE rows shift +7 and §6/§7 rows shift +7
SCORE_UPDATES = {
    (31,'D'): '=IF(C127="",C126,C127)',        # CE:  C120→C127, C119→C126
    (32,'D'): '=IF(C142="",C141,C142)',        # Doc: C135→C142, C134→C141
    (33,'E'): ('=IF(OR(C178="No",C179="No",C180="No",'
               'COUNTIF(C182:C192,"Yes")>0,COUNTIF(C157:C172,"Yes")>0)'
               ',0,SUM(E26:E32))'),
    (33,'F'): ('=IF(E33="","",IF(OR(C178="No",C179="No",C180="No",'
               'COUNTIF(C182:C192,"Yes")>0,COUNTIF(C157:C172,"Yes")>0)'
               ',"AUTO-ZERO (0%)",IF(E33>=0.85,"MEETS / EXCEEDS",'
               'IF(E33>=0.7,"DEVELOPING","NEEDS IMPROVEMENT"))))'),
}

def remap_sqref(s):
    def sub(m):
        col, row = m.group(1), int(m.group(2))
        nr = map_row(row)
        return f"{col}{nr}" if nr is not None else m.group(0)
    return re.sub(r'([A-Z]+)(\d+)', sub, s)

# ── Load ──────────────────────────────────────────────────────────
print("Loading...")
wb = load_workbook(FILE)
ws = wb[SHEET]

# ── Read reference styles (before clearing) ──────────────────────
def cell_style(r, c):
    cell = ws.cell(r, c)
    if not cell.has_style: return {}
    return {k: copy(getattr(cell, k)) for k in
            ('font','fill','border','alignment','protection',
             'number_format')}

ref_esc_header  = cell_style(96, 2)   # "5. ESCALATION JUDGMENT" heading style
ref_note        = cell_style(107, 2)  # context note style
ref_coach_hdr   = cell_style(196, 2)  # "Improvement opportunities" heading style
ref_coach_item  = cell_style(197, 3)  # item "1." style in coaching

# ── Snapshot rows 108-208 ─────────────────────────────────────────
print("Snapshotting rows 108-208...")
snap = {}
for row in ws.iter_rows(min_row=108, max_row=208):
    for cell in row:
        if cell.value is not None or cell.has_style:
            snap[(cell.row, cell.column)] = {
                'value': cell.value,
                'has_style': cell.has_style,
                **({k: copy(getattr(cell, k))
                    for k in ('font','fill','border','alignment',
                               'protection','number_format')}
                   if cell.has_style else {})
            }

row_heights = {rd.index: rd.height
               for rd in ws.row_dimensions.values() if rd.height}
merge_data  = [(m.min_row, m.min_col, m.max_row, m.max_col)
               for m in ws.merged_cells.ranges]
old_dvs     = list(ws.data_validations.dataValidation)

# ── Unmerge all; clear rows 108-219 ──────────────────────────────
print("Clearing...")
for m in list(ws.merged_cells.ranges):
    ws.unmerge_cells(str(m))
for row in ws.iter_rows(min_row=108, max_row=219):
    for cell in row:
        cell.value = None
ws.data_validations.dataValidation = []

# ── Write shifted snapshot (skip row 108) ────────────────────────
print("Writing shifted content...")
for (old_r, col), data in sorted(snap.items()):
    if old_r == 108: continue          # overwritten by new block
    d = get_delta(old_r)
    if d is None: continue
    nr = old_r + d
    cl = get_column_letter(col)
    dst = ws.cell(nr, col)
    val = data['value']
    if isinstance(val, str) and val.startswith('=') and d != 0:
        try:
            val = Translator(val, origin=f"{cl}{old_r}").translate_formula(
                row_delta=d, col_delta=0)
        except Exception as e:
            print(f"  translate err row {old_r} col {col}: {e}")
    dst.value = val
    if data.get('has_style'):
        for attr in ('font','fill','border','alignment',
                     'protection','number_format'):
            if attr in data:
                setattr(dst, attr, data[attr])

# ── Update Score Summary cross-block formulas ────────────────────
print("Patching Score Summary formulas...")
for (r, cl), formula in SCORE_UPDATES.items():
    ws.cell(r, ord(cl)-64).value = formula
    print(f"  {cl}{r} = {formula[:70]}")

# ── Helper: apply reference style ────────────────────────────────
def apply_style(cell, ref):
    for k, v in ref.items():
        try: setattr(cell, k, copy(v))
        except Exception: pass

# ── Block 1: Escalation Follow-Up (rows 108-114) ─────────────────
print("Writing Escalation Follow-Up block...")

# Row 108 – section header
c = ws.cell(108, 2)
c.value = ("ESCALATION FOLLOW-UP   ·   "
           "Complete only if an escalation was submitted")
apply_style(c, ref_esc_header)
ws.merge_cells(start_row=108, start_column=2, end_row=108, end_column=6)
ws.row_dimensions[108].height = 20

# Row 109 – label (merged through row 110)
c = ws.cell(109, 2)
c.value = "What did the receiving agent / Level 2 do to resolve or advance this case?"
apply_style(c, ref_note)
ws.merge_cells(start_row=109, start_column=2, end_row=110, end_column=3)
ws.row_dimensions[109].height = 20
ws.row_dimensions[110].height = 45   # tall input area

# Row 111 – L1 learning points sub-header
c = ws.cell(111, 2)
c.value = "L1 learning points from this escalation outcome"
apply_style(c, ref_esc_header)
ws.merge_cells(start_row=111, start_column=2, end_row=111, end_column=6)
ws.row_dimensions[111].height = 18

# Rows 112-114 – numbered items
for i, r in enumerate([112, 113, 114], 1):
    lbl = ws.cell(r, 2)
    lbl.value = f"{i}."
    apply_style(lbl, ref_note)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    ws.row_dimensions[r].height = 28

# ── Block 2: Next Session Focus (rows 209-212) ───────────────────
print("Writing Next Session Focus block...")

# Row 209 – header
c = ws.cell(209, 2)
c.value = ("NEXT SESSION FOCUS   ·   "
           "Specific behaviors to practice before the next review")
apply_style(c, ref_coach_hdr)
ws.merge_cells(start_row=209, start_column=2, end_row=209, end_column=6)
ws.row_dimensions[209].height = 20

# Rows 210-212 – numbered action items
for i, r in enumerate([210, 211, 212], 1):
    lbl = ws.cell(r, 2)
    lbl.value = f"{i}."
    apply_style(lbl, ref_coach_item)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    ws.row_dimensions[r].height = 28

# ── Rebuild merges ────────────────────────────────────────────────
print("Rebuilding merges...")
for (min_r, min_c, max_r, max_c) in merge_data:
    nm = map_row(min_r)
    nx = map_row(max_r)
    if nm is not None and nx is not None:
        try:
            ws.merge_cells(start_row=nm, start_column=min_c,
                           end_row=nx,   end_column=max_c)
        except Exception as e:
            print(f"  merge warn {min_r}-{max_r}: {e}")

# ── Rebuild row heights ───────────────────────────────────────────
for old_r, h in row_heights.items():
    nr = map_row(old_r) if old_r >= 4 else old_r
    if nr is not None:
        ws.row_dimensions[nr].height = h

# ── Rebuild DVs ───────────────────────────────────────────────────
print("Rebuilding DVs...")
for dv in old_dvs:
    nsq = remap_sqref(str(dv.sqref))
    ndv = DataValidation(
        type=dv.type, operator=dv.operator,
        formula1=dv.formula1, formula2=dv.formula2,
        allow_blank=dv.allow_blank,
        showInputMessage=dv.showInputMessage,
        showErrorMessage=dv.showErrorMessage,
        errorTitle=dv.errorTitle, error=dv.error,
        promptTitle=dv.promptTitle, prompt=dv.prompt,
    )
    ndv.sqref = nsq
    ws.add_data_validation(ndv)
    print(f"  {str(dv.sqref)[:55]!r} → {nsq[:55]!r}")

print("\nSaving...")
wb.save(FILE)
print("Done.")
