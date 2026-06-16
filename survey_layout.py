
import openpyxl
wb = openpyxl.load_workbook(r"d:\Visual Studio Code\QA Rubric\QA Scoring Form Template.xlsx")
ws = wb["Scoring Form"]

print(f"Max row: {ws.max_row}")
print()

# Print non-empty rows with their content
for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
    for cell in row:
        if cell.value is not None:
            val = str(cell.value)[:80]
            print(f"Row {cell.row:3d} Col {cell.column:2d} ({cell.coordinate}): {val}")
            break  # only show first non-empty cell per row to keep output manageable
