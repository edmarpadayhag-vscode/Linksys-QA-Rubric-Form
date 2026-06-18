
"""
Clear all user-entered data from Case Tracker, leaving formulas intact.

Touch log (rows 5-49):  clear A B C D E F G H I J K N O P Q
Keep:                   L (Weighted formula), M (Result formula)

Journey rows (56-59):   clear A E F I
Keep:                   C (Touches), D (Avg Touch), G (Team Score), H (Band)
"""

from openpyxl import load_workbook

FILE  = r"d:\Visual Studio Code\QA Rubric\QA Scoring Form Template.xlsx"
SHEET = "Case Tracker"

wb = load_workbook(FILE)
ws = wb[SHEET]

# Columns to clear in touch log rows (by number)
# A=1 B=2 C=3 D=4 E=5 F=6 G=7 H=8 I=9 J=10 K=11  N=14 O=15 P=16 Q=17
TOUCH_CLEAR_COLS = {1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 14, 15, 16, 17}

cleared = 0
for r in range(5, 50):
    for c in TOUCH_CLEAR_COLS:
        cell = ws.cell(r, c)
        if cell.value is not None:
            cell.value = None
            cleared += 1

print(f"Touch log: cleared {cleared} cells across rows 5-49")

# Journey rows: clear A (Case ID), E (Outcome), F (Continuity), I (notes)
JOURNEY_CLEAR_COLS = {1, 5, 6, 9}
cleared2 = 0
for r in range(56, 60):
    for c in JOURNEY_CLEAR_COLS:
        cell = ws.cell(r, c)
        if cell.value is not None:
            cell.value = None
            cleared2 += 1

print(f"Journey rows: cleared {cleared2} cells across rows 56-59")
print(f"Total cleared: {cleared + cleared2}")

wb.save(FILE)
print("Done.")
