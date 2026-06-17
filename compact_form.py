
"""
Remove all blank rows from Scoring Form, keeping only functional blank
input rows (Q1 answer, coaching rec, org insights, multi-touch, final
assessment, and the Escalation Follow-Up extended input area).
Rows 216/218 are orphaned duplicate headers — also dropped.
All formulas, merges, DVs, and row heights are remapped automatically.
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy
import re

FILE  = r"d:\Visual Studio Code\QA Rubric\QA Scoring Form Template.xlsx"
SHEET = "Scoring Form"

# ── Rows to drop ──────────────────────────────────────────────────
# Blank section separators, merged-away indicator gaps, orphaned rows.
# Kept blank rows (functional input areas):
#   110  – Escalation Follow-Up "What did L2 do?" extended input
#   192  – §8 Q1 answer
#   203  – Coaching recommendation input
#   210  – Org insights input
#   212  – Multi-touch notes input
#   214  – Final QA assessment input
REMOVE = {
    3,                          # header spacer
    22, 23,                     # end of §1 / before Score Summary
    34,                         # between Score Summary and §3
    43,                         # Resolution — merged indicator gap
    48,                         # between Resolution and Technical
    55, 57, 58,                 # Technical — merged indicator gaps
    63,                         # between Technical and Communication
    71,                         # Communication — merged indicator gap
    79,                         # between Communication and Ownership
    84, 86, 90,                 # Ownership — merged indicator gaps
    95,                         # between Ownership and Escalation
    102,                        # Escalation — merged indicator gap
    115,                        # between Escalation Follow-Up and CE
    122,                        # CE — merged indicator gap
    129,                        # between CE and Documentation
    137,                        # Documentation — merged indicator gap
    144, 145,                   # end of §3 / before §4
    150,                        # between §4 and §5
    154,                        # between §5 and §6
    169,                        # between §6 and §7
    189,                        # between §7 and §8
    215, 216, 217, 218, 219,    # orphaned / duplicate rows at end
}

# ── Build row map ─────────────────────────────────────────────────
row_map = {}
new_r = 0
for old_r in range(1, 220):
    if old_r not in REMOVE:
        new_r += 1
        row_map[old_r] = new_r
MAX_ROW = new_r
print(f"Rows kept: {MAX_ROW}  |  Rows removed: {len(REMOVE)}")

# ── Formula translator ────────────────────────────────────────────
def translate_formula(formula):
    def sub(m):
        col, row = m.group(1), int(m.group(2))
        nr = row_map.get(row)
        if nr is None:
            # removed row — use the last kept row before it
            for r in range(row - 1, 0, -1):
                if r in row_map:
                    nr = row_map[r]
                    break
        return f"{col}{nr}" if nr is not None else m.group(0)
    return re.sub(r'([A-Z]+)(\d+)', sub, formula)

def remap_sqref(s):
    def sub(m):
        col, row = m.group(1), int(m.group(2))
        nr = row_map.get(row)
        return f"{col}{nr}" if nr is not None else ''
    result = re.sub(r'([A-Z]+)(\d+)', sub, s)
    return re.sub(r'\s+', ' ', result).strip()

# ── Load ──────────────────────────────────────────────────────────
print("Loading...")
wb = load_workbook(FILE)
ws = wb[SHEET]

# ── Snapshot everything ───────────────────────────────────────────
print("Snapshotting...")
snap = {}
for row in ws.iter_rows(min_row=1, max_row=219):
    for cell in row:
        if cell.value is not None or cell.has_style:
            snap[(cell.row, cell.column)] = {
                'value': cell.value,
                'has_style': cell.has_style,
                **({k: copy(getattr(cell, k))
                    for k in ('font','fill','border','alignment',
                              'protection','number_format')}
                   if cell.has_style else {})
            }

row_heights = {rd.index: rd.height
               for rd in ws.row_dimensions.values() if rd.height}
merge_data  = [(m.min_row, m.min_col, m.max_row, m.max_col)
               for m in ws.merged_cells.ranges]
old_dvs     = list(ws.data_validations.dataValidation)

# ── Clear ─────────────────────────────────────────────────────────
print("Clearing...")
for m in list(ws.merged_cells.ranges):
    ws.unmerge_cells(str(m))
for row in ws.iter_rows(min_row=1, max_row=220):
    for cell in row:
        cell.value = None
ws.data_validations.dataValidation = []

# ── Write at new positions ────────────────────────────────────────
print("Writing cells...")
for (old_r, col), data in sorted(snap.items()):
    if old_r in REMOVE:
        continue
    nr = row_map.get(old_r)
    if nr is None:
        continue
    dst = ws.cell(nr, col)
    val = data['value']
    if isinstance(val, str) and val.startswith('='):
        val = translate_formula(val)
    dst.value = val
    if data.get('has_style'):
        for attr in ('font','fill','border','alignment',
                     'protection','number_format'):
            if attr in data:
                setattr(dst, attr, data[attr])

# ── Rebuild merges ────────────────────────────────────────────────
print("Rebuilding merges...")
for (min_r, min_c, max_r, max_c) in merge_data:
    nm = row_map.get(min_r)
    nx = row_map.get(max_r)
    # If max_r was a removed row, extend to last kept row before it
    if nm is not None and nx is None:
        for r in range(max_r - 1, min_r - 1, -1):
            if r in row_map:
                nx = row_map[r]
                break
    if nm is not None and nx is not None:
        try:
            ws.merge_cells(start_row=nm, start_column=min_c,
                           end_row=nx,   end_column=max_c)
        except Exception as e:
            print(f"  merge warn {min_r}-{max_r}: {e}")

# ── Rebuild row heights ───────────────────────────────────────────
for old_r, h in row_heights.items():
    nr = row_map.get(old_r)
    if nr is not None:
        ws.row_dimensions[nr].height = h

# ── Rebuild DVs ───────────────────────────────────────────────────
print("Rebuilding DVs...")
for dv in old_dvs:
    new_sqref = remap_sqref(str(dv.sqref))
    if not new_sqref.strip():
        continue
    ndv = DataValidation(
        type=dv.type, operator=dv.operator,
        formula1=dv.formula1, formula2=dv.formula2,
        allow_blank=dv.allow_blank,
        showInputMessage=dv.showInputMessage,
        showErrorMessage=dv.showErrorMessage,
        errorTitle=dv.errorTitle, error=dv.error,
        promptTitle=dv.promptTitle, prompt=dv.prompt,
    )
    ndv.sqref = new_sqref
    ws.add_data_validation(ndv)
    print(f"  {str(dv.sqref)[:55]!r} → {new_sqref[:55]!r}")

print(f"\nFinal row count: {MAX_ROW}")
print("Saving...")
wb.save(FILE)
print("Done.")
