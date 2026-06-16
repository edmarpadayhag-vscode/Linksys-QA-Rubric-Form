
import openpyxl
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy
import re

FILE = r"d:\Visual Studio Code\QA Rubric\QA Scoring Form Template.xlsx"
SHEET = "Scoring Form"

# (old_start, old_end, new_start)  — new_start - old_start = delta
BLOCKS = [
    (4,   22,  4),    # Case Info          delta=0
    (127, 136, 24),   # Score Summary      delta=-103
    (24,  126, 35),   # Category Assess.   delta=+11
    (138, 142, 139),  # Resolution         delta=+1
    (143, 146, 144),  # Disconnected       delta=+1
    (147, 165, 148),  # Critical Failures  delta=+1
    (189, 207, 168),  # Auto-Zero          delta=-21
    (167, 187, 188),  # Coaching           delta=+21
]

def get_delta(row):
    for (os, oe, ns) in BLOCKS:
        if os <= row <= oe:
            return ns - os
    return None

def map_row(row):
    if row <= 3:
        return row
    d = get_delta(row)
    return row + d if d is not None else None

# Cross-block formulas: key=(new_row, col_letter), value=corrected formula string
CROSS_BLOCK = {
    (26,'D'): '=IF(C46="",C45,C46)',
    (27,'D'): '=IF(C61="",C60,C61)',
    (28,'D'): '=IF(C77="",C76,C77)',
    (29,'D'): '=IF(C93="",C92,C93)',
    (30,'D'): '=IF(C106="",C105,C106)',
    (31,'D'): '=IF(C120="",C119,C120)',
    (32,'D'): '=IF(C135="",C134,C135)',
    (33,'E'): ('=IF(OR(C171="No",C172="No",C173="No",'
               'COUNTIF(C175:C185,"Yes")>0,COUNTIF(C150:C165,"Yes")>0),0,SUM(E26:E32))'),
    (33,'F'): ('=IF(E33="","",IF(OR(C171="No",C172="No",C173="No",'
               'COUNTIF(C175:C185,"Yes")>0,COUNTIF(C150:C165,"Yes")>0),"AUTO-ZERO (0%)",'
               'IF(E33>=0.85,"MEETS / EXCEEDS",IF(E33>=0.7,"DEVELOPING","NEEDS IMPROVEMENT"))))'),
    (146,'E'): ('=IF(AND(C145="Yes",OR(C146="None",C146="N.A. — no disconnect")),'
                '"⚠ MISSED OWNERSHIP OPPORTUNITY — flag in Customer Ownership and Critical Failures","OK")'),
    (166,'C'): ('=IF(COUNTIF(C150:C165,"Yes")>0,"YES — "&COUNTIF(C150:C165,"Yes")&'
                '" condition(s) flagged — automatic review required regardless of weighted score",'
                '"No critical failures")'),
    (186,'C'): ('=IF(OR(C171="No",C172="No",C173="No",'
                'COUNTIF(C175:C185,"Yes")>0,COUNTIF(C150:C165,"Yes")>0),'
                '"YES — Quality Score forced to 0%","No")'),
}

# Section header text replacements at their NEW row positions
SECTION_RENAMES = {
    24:  ('3 · SCORE SUMMARY',     '2 · SCORE SUMMARY'),
    35:  ('2 · CATEGORY',          '3 · CATEGORY'),
    168: ('8 · AUTO-ZERO',         '7 · AUTO-ZERO'),
    188: ('7 · COACHING',          '8 · COACHING'),
}

def remap_sqref(s):
    """Remap all row numbers in a sqref string using map_row()."""
    def replace_ref(m):
        col, row = m.group(1), int(m.group(2))
        nr = map_row(row)
        return f"{col}{nr}" if nr is not None else m.group(0)
    return re.sub(r'([A-Z]+)(\d+)', replace_ref, s)

# ── Load ──────────────────────────────────────────────────────────
print("Loading workbook...")
wb = load_workbook(FILE)
ws = wb[SHEET]

# ── 1. Snapshot rows 24-207 ────────────────────────────────────────
print("Reading cell data...")
snapshot = {}  # (old_row, col_int) → dict
for row in ws.iter_rows(min_row=24, max_row=207):
    for cell in row:
        if cell.value is not None or cell.has_style:
            snapshot[(cell.row, cell.column)] = {
                'value': cell.value,
                'has_style': cell.has_style,
                'font':        copy(cell.font)       if cell.has_style else None,
                'fill':        copy(cell.fill)       if cell.has_style else None,
                'border':      copy(cell.border)     if cell.has_style else None,
                'alignment':   copy(cell.alignment)  if cell.has_style else None,
                'number_format': cell.number_format,
                'protection':  copy(cell.protection) if cell.has_style else None,
            }

# ── 2. Read row heights ────────────────────────────────────────────
row_heights = {rd.index: rd.height for rd in ws.row_dimensions.values() if rd.height}

# ── 3. Read merges ─────────────────────────────────────────────────
merge_data = [(m.min_row, m.min_col, m.max_row, m.max_col)
              for m in ws.merged_cells.ranges]

# ── 4. Read DVs ────────────────────────────────────────────────────
old_dvs = list(ws.data_validations.dataValidation)

# ── 5. Unmerge all; clear rows 24-210 ─────────────────────────────
print("Clearing...")
for m in list(ws.merged_cells.ranges):
    ws.unmerge_cells(str(m))

for row in ws.iter_rows(min_row=24, max_row=210):
    for cell in row:
        cell.value = None

ws.data_validations.dataValidation = []

# ── 6. Write cells at new positions ───────────────────────────────
print("Writing cells...")
errors = []
for (old_r, col), data in sorted(snapshot.items()):
    delta = get_delta(old_r)
    if delta is None:
        continue   # skip empty/gap rows in the old layout
    nr = old_r + delta

    col_letter = get_column_letter(col)
    dst = ws.cell(nr, col)

    val = data['value']

    # Formula translation
    if isinstance(val, str) and val.startswith('='):
        key = (nr, col_letter)
        if key in CROSS_BLOCK:
            val = CROSS_BLOCK[key]
        elif delta != 0:
            origin = f"{col_letter}{old_r}"
            try:
                val = Translator(val, origin=origin).translate_formula(
                    row_delta=delta, col_delta=0)
            except Exception as e:
                errors.append(f"  translate error ({old_r},{col}): {e}")

    dst.value = val

    if data['has_style']:
        dst.font          = data['font']
        dst.fill          = data['fill']
        dst.border        = data['border']
        dst.alignment     = data['alignment']
        dst.number_format = data['number_format']
        dst.protection    = data['protection']

if errors:
    for e in errors:
        print(e)

# ── 7. Section header renaming ─────────────────────────────────────
print("Renaming headers...")
for new_r, (old_frag, new_frag) in SECTION_RENAMES.items():
    cell = ws.cell(new_r, 2)
    if cell.value and old_frag in str(cell.value):
        cell.value = str(cell.value).replace(old_frag, new_frag)
        print(f"  {new_r}: '{old_frag}' → '{new_frag}'")
    else:
        print(f"  WARNING row {new_r}: value={repr(cell.value)}")

# ── 8. Rebuild merges ──────────────────────────────────────────────
print("Rebuilding merges...")
for (min_r, min_c, max_r, max_c) in merge_data:
    new_min = map_row(min_r)
    new_max = map_row(max_r)
    if new_min is not None and new_max is not None:
        ws.merge_cells(start_row=new_min, start_column=min_c,
                       end_row=new_max,   end_column=max_c)
    else:
        print(f"  WARNING: skipped merge {min_r},{min_c}-{max_r},{max_c}")

# ── 9. Rebuild row heights ─────────────────────────────────────────
print("Applying row heights...")
for old_r, height in row_heights.items():
    nr = map_row(old_r) if old_r >= 4 else old_r
    if nr is not None:
        ws.row_dimensions[nr].height = height

# ── 10. Rebuild data validations ───────────────────────────────────
print("Rebuilding DVs...")
for dv in old_dvs:
    new_sqref = remap_sqref(str(dv.sqref))
    new_dv = DataValidation(
        type=dv.type,
        operator=dv.operator,
        formula1=dv.formula1,
        formula2=dv.formula2,
        allow_blank=dv.allow_blank,
        showInputMessage=dv.showInputMessage,
        showErrorMessage=dv.showErrorMessage,
        errorTitle=dv.errorTitle,
        error=dv.error,
        promptTitle=dv.promptTitle,
        prompt=dv.prompt,
    )
    new_dv.sqref = new_sqref
    ws.add_data_validation(new_dv)
    print(f"  {str(dv.sqref)[:60]!r}  →  {new_sqref[:60]!r}")

# ── Save ──────────────────────────────────────────────────────────
print("\nSaving...")
wb.save(FILE)
print("Done.")
