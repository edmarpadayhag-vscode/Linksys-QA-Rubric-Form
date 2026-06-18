
"""Remove blank row 75 from Scoring Form. Rows 76-189 shift to 75-188."""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy
import re

FILE  = r"d:\Visual Studio Code\QA Rubric\QA Scoring Form Template.xlsx"
SHEET = "Scoring Form"

REMOVE_ROW = 75
OLD_MAX    = 189

def map_row(r):
    if r < REMOVE_ROW:  return r
    if r == REMOVE_ROW: return None
    return r - 1

def translate_formula(formula):
    def sub(m):
        col, row = m.group(1), int(m.group(2))
        nr = map_row(row)
        return f"{col}{nr}" if nr is not None else m.group(0)
    return re.sub(r'([A-Z]+)(\d+)', sub, formula)

def remap_sqref(s):
    def sub(m):
        col, row = m.group(1), int(m.group(2))
        nr = map_row(row)
        return f"{col}{nr}" if nr is not None else ''
    result = re.sub(r'([A-Z]+)(\d+)', sub, s)
    return re.sub(r'\s+', ' ', result).strip()

wb = load_workbook(FILE)
ws = wb[SHEET]

# Snapshot rows 76-189 (everything that shifts)
print("Snapshotting...")
snap = {}
for row in ws.iter_rows(min_row=REMOVE_ROW + 1, max_row=OLD_MAX):
    for cell in row:
        if cell.value is not None or cell.has_style:
            snap[(cell.row, cell.column)] = {
                'value': cell.value,
                'has_style': cell.has_style,
                **({k: copy(getattr(cell, k))
                    for k in ('font','fill','border','alignment',
                              'protection','number_format')}
                   if cell.has_style else {})
            }

row_heights = {rd.index: rd.height for rd in ws.row_dimensions.values() if rd.height}
merge_data  = [(m.min_row, m.min_col, m.max_row, m.max_col)
               for m in ws.merged_cells.ranges]
old_dvs     = list(ws.data_validations.dataValidation)
old_cf      = [(str(k.sqref) if hasattr(k, 'sqref') else str(k), list(v))
               for k, v in ws.conditional_formatting._cf_rules.items()]

# Clear row 75 onward
print("Clearing...")
for m in list(ws.merged_cells.ranges):
    ws.unmerge_cells(str(m))
for row in ws.iter_rows(min_row=REMOVE_ROW, max_row=OLD_MAX):
    for cell in row:
        cell.value = None
ws.data_validations.dataValidation = []

# Write shifted content
print("Writing shifted rows...")
for (old_r, col), data in sorted(snap.items()):
    nr = map_row(old_r)
    if nr is None:
        continue
    dst = ws.cell(nr, col)
    val = data['value']
    if isinstance(val, str) and val.startswith('='):
        val = translate_formula(val)
    dst.value = val
    if data.get('has_style'):
        for attr in ('font','fill','border','alignment','protection','number_format'):
            if attr in data:
                setattr(dst, attr, data[attr])

# Rebuild merges
print("Rebuilding merges...")
for (min_r, min_c, max_r, max_c) in merge_data:
    nm = map_row(min_r)
    nx = map_row(max_r)
    if nm is None or nx is None:
        continue
    try:
        ws.merge_cells(start_row=nm, start_column=min_c,
                       end_row=nx,   end_column=max_c)
    except Exception as e:
        print(f"  merge warn {min_r}-{max_r}: {e}")

# Rebuild row heights
for old_r, h in row_heights.items():
    nr = map_row(old_r) if old_r >= REMOVE_ROW else old_r
    if nr is not None:
        ws.row_dimensions[nr].height = h

# Rebuild DVs
print("Rebuilding DVs...")
for dv in old_dvs:
    new_sq = remap_sqref(str(dv.sqref))
    if not new_sq.strip():
        continue
    ndv = DataValidation(
        type=dv.type, operator=dv.operator,
        formula1=dv.formula1, formula2=dv.formula2,
        allow_blank=dv.allow_blank,
        showInputMessage=dv.showInputMessage,
        showErrorMessage=dv.showErrorMessage,
        errorTitle=dv.errorTitle, error=dv.error,
        promptTitle=dv.promptTitle, prompt=dv.prompt,
    )
    ndv.sqref = new_sq
    ws.add_data_validation(ndv)
    if new_sq != str(dv.sqref):
        print(f"  {str(dv.sqref)[:55]!r} → {new_sq[:55]!r}")

# Rebuild CF
print("Rebuilding conditional formatting...")
ws.conditional_formatting._cf_rules = {}
for cf_range_str, rules in old_cf:
    new_range_str = remap_sqref(cf_range_str)
    if new_range_str:
        for rule in rules:
            ws.conditional_formatting.add(new_range_str, rule)

# Ghost cell cleanup
new_max = OLD_MAX - 1
to_del = [(r, c) for (r, c) in list(ws._cells.keys()) if r > new_max]
for key in to_del:
    del ws._cells[key]
print(f"  Ghost cells removed: {len(to_del)}  (new max row = {new_max})")

# Quick verify
print("\n=== Ownership rows 72-80 after shift ===")
for r in range(72, 81):
    v = ws.cell(r, 2).value
    print(f"  B{r}: {repr(str(v)[:70]) if v else '(blank)'}")

wb.save(FILE)
print("\nDone.")
