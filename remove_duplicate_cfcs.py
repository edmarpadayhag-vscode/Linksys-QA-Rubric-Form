
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy
import re

FILE  = r"d:\Visual Studio Code\QA Rubric\QA Scoring Form Template.xlsx"
SHEET = "Scoring Form"

# Rows to REMOVE from §6 (duplicated in §8 Auto-Zero)
REMOVE = {165, 168, 169, 171, 172}

def map_row(row):
    """Map current rows to new positions after removing 5 §6 conditions."""
    if row <= 164:   return row
    if row in REMOVE: return None      # deleted
    if row == 166:   return 165        # Case marked Resolved
    if row == 167:   return 166        # Customer complaint/legal
    if row == 170:   return 167        # Survey/CSAT manipulation
    if row >= 173:   return row - 5    # §7, §8 and everything below shift -5
    return None

# Hardcoded cross-section formulas (row references change non-trivially)
HARDCODE = {
    # C168 — CRITICAL FAILURE FLAGGED? (was C173, range shrinks from 157:172 → 157:167)
    (168,'C'): ('=IF(COUNTIF(C157:C167,"Yes")>0,"YES — "&COUNTIF(C157:C167,"Yes")&'
                '" condition(s) flagged — automatic review required regardless of weighted score",'
                '"No critical failures")'),
    # E33 & F33 — §7 gates/list shift -5; §6 range shrinks
    (33,'E'):  ('=IF(OR(C173="No",C174="No",C175="No",'
                'COUNTIF(C177:C187,"Yes")>0,COUNTIF(C157:C167,"Yes")>0),0,SUM(E26:E32))'),
    (33,'F'):  ('=IF(E33="","",IF(OR(C173="No",C174="No",C175="No",'
                'COUNTIF(C177:C187,"Yes")>0,COUNTIF(C157:C167,"Yes")>0),"AUTO-ZERO (0%)",'
                'IF(E33>=0.85,"MEETS / EXCEEDS",IF(E33>=0.7,"DEVELOPING","NEEDS IMPROVEMENT"))))'),
    # C188 — AUTO-ZERO TRIGGERED? (was C193, same ref updates)
    (188,'C'): ('=IF(OR(C173="No",C174="No",C175="No",'
                'COUNTIF(C177:C187,"Yes")>0,COUNTIF(C157:C167,"Yes")>0),'
                '"YES — Quality Score forced to 0%","No")'),
}

def remap_sqref(s):
    """Remap individual cell refs; for ranges remap start and end separately."""
    def sub(m):
        col, row = m.group(1), int(m.group(2))
        nr = map_row(row)
        # Removed rows: drop from sqref (handle below for ranges)
        return f"{col}{nr}" if nr is not None else f"{col}REMOVED"
    return re.sub(r'([A-Z]+)(\d+)', sub, s)

# ── Load ──────────────────────────────────────────────────────────
print("Loading...")
wb = load_workbook(FILE)
ws = wb[SHEET]

# ── Snapshot rows 165-219 ─────────────────────────────────────────
print("Snapshotting rows 165-219...")
snap = {}
for row in ws.iter_rows(min_row=165, max_row=219):
    for cell in row:
        if cell.value is not None or cell.has_style:
            snap[(cell.row, cell.column)] = {
                'value': cell.value,
                'has_style': cell.has_style,
                **({k: copy(getattr(cell, k))
                    for k in ('font','fill','border','alignment',
                              'protection','number_format')}
                   if cell.has_style else {})
            }

row_heights = {rd.index: rd.height
               for rd in ws.row_dimensions.values() if rd.height}
merge_data  = [(m.min_row, m.min_col, m.max_row, m.max_col)
               for m in ws.merged_cells.ranges]
old_dvs     = list(ws.data_validations.dataValidation)

# ── Unmerge all; clear rows 165-214 ──────────────────────────────
print("Clearing...")
for m in list(ws.merged_cells.ranges):
    ws.unmerge_cells(str(m))
for row in ws.iter_rows(min_row=165, max_row=214):
    for cell in row:
        cell.value = None
ws.data_validations.dataValidation = []

# ── Write shifted snapshot (skip removed rows) ───────────────────
print("Writing shifted content...")
for (old_r, col), data in sorted(snap.items()):
    nr = map_row(old_r)
    if nr is None:
        continue    # removed row
    cl  = get_column_letter(col)
    dst = ws.cell(nr, col)
    val = data['value']
    # Check hardcode table first (cross-section formulas)
    if (nr, cl) in HARDCODE:
        val = HARDCODE[(nr, cl)]
    dst.value = val
    if data.get('has_style'):
        for attr in ('font','fill','border','alignment',
                     'protection','number_format'):
            if attr in data:
                setattr(dst, attr, data[attr])

# ── Apply remaining hardcoded formulas (E33, F33 stay in place) ──
print("Patching Score Summary formulas...")
for (r, cl), formula in HARDCODE.items():
    if r == 33:   # E33/F33 are outside the snapshot range, patch directly
        ws.cell(r, ord(cl)-64).value = formula
        print(f"  {cl}{r}: {formula[:70]}")

# ── Rebuild merges ────────────────────────────────────────────────
print("Rebuilding merges...")
for (min_r, min_c, max_r, max_c) in merge_data:
    nm = map_row(min_r)
    nx = map_row(max_r)
    if nm is not None and nx is not None:
        try:
            ws.merge_cells(start_row=nm, start_column=min_c,
                           end_row=nx,   end_column=max_c)
        except Exception as e:
            print(f"  merge warn {min_r}-{max_r}: {e}")

# ── Rebuild row heights ───────────────────────────────────────────
for old_r, h in row_heights.items():
    nr = map_row(old_r) if old_r >= 165 else old_r
    if nr is not None:
        ws.row_dimensions[nr].height = h

# ── Rebuild DVs with corrected sqrefs ─────────────────────────────
print("Rebuilding DVs...")
for dv in old_dvs:
    raw = str(dv.sqref)

    # §6 conditions range: C157:C172 → C157:C167  (range shrinks)
    raw = raw.replace('C157:C172', 'C157:C167')
    # §7 A-K list:         C182:C192 → C177:C187  (shift -5)
    raw = raw.replace('C182:C192', 'C177:C187')
    # §7 gates:            C178:C180 → C173:C175  (shift -5)
    raw = raw.replace('C178:C180', 'C173:C175')

    # All DV cells outside the three explicit ranges are at rows ≤ 164 (unchanged)
    # or E10 — no further remapping needed.
    new_sqref = raw

    ndv = DataValidation(
        type=dv.type, operator=dv.operator,
        formula1=dv.formula1, formula2=dv.formula2,
        allow_blank=dv.allow_blank,
        showInputMessage=dv.showInputMessage,
        showErrorMessage=dv.showErrorMessage,
        errorTitle=dv.errorTitle, error=dv.error,
        promptTitle=dv.promptTitle, prompt=dv.prompt,
    )
    ndv.sqref = new_sqref
    ws.add_data_validation(ndv)
    if raw != str(dv.sqref):
        print(f"  {str(dv.sqref)[:55]!r} → {new_sqref[:55]!r}")

print("\nSaving...")
wb.save(FILE)
print("Done.")
