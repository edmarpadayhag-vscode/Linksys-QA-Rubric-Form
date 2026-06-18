
"""
Implement Option A — Journey Outcome & Continuity scoring anchors.

1. Case Tracker  — split E56:F59 DV into two with input message tooltips
2. Scoring Guide — add new section after row 143 with full anchor tables
3. Multi-Touch Guide — update formula text (C37) and expand anchors (C38/C39)
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy

FILE = r"d:\Visual Studio Code\QA Rubric\QA Scoring Form Template.xlsx"
wb   = load_workbook(FILE)

# ── Shared style helpers ───────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def fnt(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic)

def aln(wrap=True, v="top", h="left"):
    return Alignment(wrap_text=wrap, vertical=v, horizontal=h)

_s = Side(style="thin", color="BFBFBF")
def bdr():
    return Border(left=_s, right=_s, top=_s, bottom=_s)

# ══════════════════════════════════════════════════════════════════
# 1. CASE TRACKER — split DV into two with tooltips
# ══════════════════════════════════════════════════════════════════
print("=== Case Tracker: updating DVs ===")
ws_ct = wb["Case Tracker"]

# Remove the combined E56:F59 DV
old_dvs = list(ws_ct.data_validations.dataValidation)
ws_ct.data_validations.dataValidation = [
    dv for dv in old_dvs
    if "E56" not in str(dv.sqref) and "F56" not in str(dv.sqref)
]
print("  Removed combined E56:F59 DV")

# Journey Outcome tooltip (≤255 chars)
jo_prompt = (
    "5 = Resolved & confirmed  "
    "4 = Strong path: RMA / escalation / limitation explained  "
    "3 = Partial — next step exists but issue open  "
    "2 = Minimal progress  "
    "1 = No meaningful progress  "
    "0 = Worsened or contradictory guidance"
)
dv_jo = DataValidation(
    type="whole", operator="between",
    formula1=0, formula2=5,
    allow_blank=True,
    showInputMessage=True,
    promptTitle="Journey Outcome (0–5)",
    prompt=jo_prompt,
    showErrorMessage=True,
    errorTitle="Out of range",
    error="Enter a whole number between 0 and 5.",
)
dv_jo.sqref = "E56:E59"
ws_ct.add_data_validation(dv_jo)
print(f"  Added Journey Outcome DV  ({len(jo_prompt)} chars)")

# Continuity tooltip (≤255 chars)
cont_prompt = (
    "5 = Full context, no repeated questions, statuses always truthful  "
    "4 = Minor gaps, no customer impact  "
    "3 = Customer re-explained parts of issue  "
    "2 = Significant context loss across touches  "
    "1 = Near-zero continuity  "
    "0 = Contradictory guidance or breakdown"
)
dv_cont = DataValidation(
    type="whole", operator="between",
    formula1=0, formula2=5,
    allow_blank=True,
    showInputMessage=True,
    promptTitle="Continuity of Support (0–5)",
    prompt=cont_prompt,
    showErrorMessage=True,
    errorTitle="Out of range",
    error="Enter a whole number between 0 and 5.",
)
dv_cont.sqref = "F56:F59"
ws_ct.add_data_validation(dv_cont)
print(f"  Added Continuity DV  ({len(cont_prompt)} chars)")

# ══════════════════════════════════════════════════════════════════
# 2. SCORING GUIDE — append Team Score section after row 143
# ══════════════════════════════════════════════════════════════════
print("\n=== Scoring Guide: appending Team Score section ===")
ws_sg = wb["Scoring Guide"]

# Copy column widths and header styles from existing rows for consistency
# Existing col widths: A≈20, B≈55, C≈12, D≈60, E≈60 (approximate)

# Content: (section, indicator, score, description, example)
JOURNEY_OUTCOME = [
    ("TEAM SCORE\nJourney Outcome\n(0–5)",
     "Journey Outcome — overall resolution quality across all touches in the case",
     "5",
     "Issue fully resolved and confirmed with the customer. Positive close with no open actions.",
     "Customer confirmed streaming issue resolved after MX4200 replacement. Ticket closed Resolved with customer acknowledgment."),
    ("", "", "4",
     "Strong resolution path achieved — proper RMA, valid escalation submitted with follow-through, or known limitation accurately explained and accepted.",
     "In-warranty hardware fault confirmed across 2 touches; RMA initiated and tracking number provided at close. Customer satisfied with next steps."),
    ("", "", "3",
     "Partial progress across the journey. Customer has a defined next step but the core issue remains open or unconfirmed at close.",
     "After 2 touches, ISP issue identified as likely cause; customer advised to contact ISP. Issue not independently verified but customer left with an actionable step."),
    ("", "", "2",
     "Minimal progress despite multiple touches. No clear resolution path at close; customer no closer to resolution than at first contact.",
     "3 touches across 2 weeks. Each agent troubleshot independently with no continuity. No resolution path defined and no escalation attempted."),
    ("", "", "1",
     "No meaningful progress. Multiple touches produced no actionable outcome; customer has no next steps at final close.",
     "4 contacts, each ending with 'monitor and call back if it happens again.' No escalation, no RMA, no root-cause identified."),
    ("", "", "0",
     "Outcome worsened. Contradictory guidance given across touches, customer actively harmed, or a critical failure occurred with no correction by the chain.",
     "Customer received conflicting advice — one agent blamed ISP, next diagnosed hardware fault — with no diagnostic data supporting either. Customer filed a formal complaint."),
]

CONTINUITY = [
    ("TEAM SCORE\nContinuity of\nSupport (0–5)",
     "Continuity of Support — how well context, ownership, and accuracy were maintained across handoffs",
     "5",
     "Full context preserved at every handoff. No agent restarted from zero. Ticket statuses truthful throughout. All handoffs documented with next steps and named owner.",
     "Each agent opened with acknowledgment of prior steps — no repeated diagnostic questions. Ticket status accurately reflected case state at every touch. Clear handoff note at each close."),
    ("", "", "4",
     "Minor context gap at one handoff with no meaningful customer impact. Most handoffs documented with next steps.",
     "Agent 2 re-asked for the model number despite it being in notes, but quickly moved on. All other handoff details intact. Customer was not asked to repeat their issue."),
    ("", "", "3",
     "Noticeable context gaps at one or more handoffs. Customer had to re-explain part of their issue. Statuses mostly accurate but handoff documentation incomplete.",
     "Customer had to repeat troubleshooting history to agent 3. Prior notes were present but not reviewed before the call. Ticket status was accurate but lacked next-step documentation."),
    ("", "", "2",
     "Significant context loss across the journey. Customer re-explained their issue on multiple contacts. Ticket statuses inconsistent or misleading at some points.",
     "Customer called 3 times. Each agent asked the same diagnostic questions. No evidence of cross-referencing prior notes. Ticket was marked Resolved at touch 2 while issue was still open."),
    ("", "", "1",
     "Near-zero continuity. Agents effectively started fresh on each contact. Ticket statuses inaccurate at close of multiple touches.",
     "Case had 5 touches. Notes were absent or generic ('called customer, no resolution'). Customer verbally expressed frustration at re-explaining their problem on every contact."),
    ("", "", "0",
     "Complete breakdown. Contradictory instructions given by different agents. Customer actively misled or harmed by lack of coordination between touches.",
     "Agent 1 instructed a factory reset and re-pairing. Agent 2, with no awareness of the reset, told customer the issue was ISP-side — effectively undoing Agent 1's work with no evidence to support the ISP claim."),
]

# Row color pattern matching the existing Scoring Guide
RATING_FILLS = {
    "5": "E2EFDA",   # green (Met equivalent)
    "4": "EBF7E3",   # light green
    "3": "FFF2CC",   # yellow (Partial equivalent)
    "2": "FCE4D6",   # salmon (Missed equivalent)
    "1": "F8D7D7",   # light red
    "0": "C00000",   # dark red
}
RATING_FONTS = {
    "0": "FFFFFF",   # white text on dark red
}

# Section header row
start_row = ws_sg.max_row + 2   # leave one blank gap

# Section banner
ws_sg.merge_cells(start_row=start_row, start_column=1,
                  end_row=start_row, end_column=5)
c = ws_sg.cell(start_row, 1)
c.value     = "TEAM SCORE — JOURNEY OUTCOME & CONTINUITY OF SUPPORT   ·   Used in Case Tracker multi-touch scoring"
c.font      = fnt(bold=True, color="FFFFFF", size=10)
c.fill      = fill("2E75B6")
c.alignment = aln(v="center", h="center")
c.border    = bdr()
ws_sg.row_dimensions[start_row].height = 20
print(f"  Section banner at row {start_row}")

# Column sub-headers
sub_row = start_row + 1
for col, txt in [(1,"Component"),(2,"Description"),(3,"Score"),(4,"What This Looks Like"),(5,"Example Scenario")]:
    c = ws_sg.cell(sub_row, col)
    c.value     = txt
    c.font      = fnt(bold=True, color="FFFFFF", size=9.5)
    c.fill      = fill("4472C4")
    c.alignment = aln(v="center", h="center")
    c.border    = bdr()
ws_sg.row_dimensions[sub_row].height = 16

data_row = sub_row + 1

def write_anchor_rows(rows_data):
    global data_row
    for (section, indicator, score, desc, example) in rows_data:
        r = data_row
        fg = RATING_FILLS.get(score, "FFFFFF")
        fc = RATING_FONTS.get(score, "000000")

        # A – section
        c = ws_sg.cell(r, 1)
        c.value     = section
        c.font      = fnt(bold=bool(section), size=9, color="2E4057")
        c.fill      = fill("F2F2F2")
        c.alignment = aln(v="top", h="left")
        c.border    = bdr()

        # B – indicator
        c = ws_sg.cell(r, 2)
        c.value     = indicator
        c.font      = fnt(size=9, italic=bool(indicator))
        c.fill      = fill("F2F2F2")
        c.alignment = aln(v="top", h="left")
        c.border    = bdr()

        # C – score
        c = ws_sg.cell(r, 3)
        c.value     = score
        c.font      = fnt(bold=True, size=9.5, color=fc)
        c.fill      = fill(fg)
        c.alignment = aln(v="center", h="center")
        c.border    = bdr()

        # D – description
        c = ws_sg.cell(r, 4)
        c.value     = desc
        c.font      = fnt(size=9)
        c.fill      = fill(fg)
        c.alignment = aln(v="top", h="left")
        c.border    = bdr()

        # E – example
        c = ws_sg.cell(r, 5)
        c.value     = example
        c.font      = fnt(size=9, italic=True)
        c.fill      = fill(fg)
        c.alignment = aln(v="top", h="left")
        c.border    = bdr()

        ws_sg.row_dimensions[r].height = 52
        data_row += 1

write_anchor_rows(JOURNEY_OUTCOME)
# Spacer
data_row += 1
write_anchor_rows(CONTINUITY)
print(f"  Anchor rows written through row {data_row - 1}")

# ══════════════════════════════════════════════════════════════════
# 3. MULTI-TOUCH GUIDE — update formula + expand anchor cells
# ══════════════════════════════════════════════════════════════════
print("\n=== Multi-Touch Guide: updating ===")
ws_mt = wb["Multi-Touch Guide"]

# C37 — update formula description
ws_mt["C37"].value = (
    "Team Score  =  50% × Average Touch Weighted Score  +  30% × (Journey Outcome ÷ 5)  +  20% × (Continuity ÷ 5)\n"
    "When a case has BOTH Level 1 and Level 2 touches, the 50% touch component splits:  "
    "20% × Level 1 avg  +  30% × Level 2 avg  (Level 2 carries higher weight in mixed cases)"
)
ws_mt["C37"].alignment = Alignment(wrap_text=True, vertical="top")
ws_mt.row_dimensions[37].height = 42
print("  Updated C37 formula description")

# C38 — Journey Outcome expanded
ws_mt["C38"].value = (
    "5  ·  Issue fully resolved and confirmed — positive close, no open actions\n"
    "4  ·  Strong path: proper RMA, valid escalation with follow-through, or limitation accurately explained\n"
    "3  ·  Partial — customer has a next step but issue remains open or unconfirmed\n"
    "2  ·  Minimal progress despite multiple touches; no resolution path at close\n"
    "1  ·  No meaningful progress; customer has no next steps at final close\n"
    "0  ·  Worsened — contradictory guidance, critical failure, or customer abandoned"
)
ws_mt["C38"].alignment = Alignment(wrap_text=True, vertical="top")
ws_mt.row_dimensions[38].height = 72
print("  Updated C38 Journey Outcome anchors")

# C39 — Continuity expanded
ws_mt["C39"].value = (
    "5  ·  Full context at every handoff; no repeated questions; statuses always truthful; next steps documented\n"
    "4  ·  Minor gap at one handoff, no customer impact; most handoffs documented\n"
    "3  ·  Customer re-explained parts of issue; statuses mostly accurate; handoff notes incomplete\n"
    "2  ·  Significant context loss; customer re-explained issue on multiple contacts; statuses inconsistent\n"
    "1  ·  Near-zero continuity; agents restarted fresh each contact; statuses inaccurate at multiple closes\n"
    "0  ·  Complete breakdown — contradictory instructions given; customer actively harmed by lack of coordination"
)
ws_mt["C39"].alignment = Alignment(wrap_text=True, vertical="top")
ws_mt.row_dimensions[39].height = 72
print("  Updated C39 Continuity anchors")

# ── Save ───────────────────────────────────────────────────────────
print("\nSaving...")
wb.save(FILE)
print("Done.")
