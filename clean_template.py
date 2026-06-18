
"""
1. Clear case-specific notes from template rows 43, 55, 69, 80, 91, 110, 123
2. Fix Row 61 Communication indicator wording
"""

from openpyxl import load_workbook

FILE  = r"d:\Visual Studio Code\QA Rubric\QA Scoring Form Template.xlsx"
SHEET = "Scoring Form"

wb = load_workbook(FILE)
ws = wb[SHEET]

# ── 1. Clear contaminated rows ─────────────────────────────────────
CONTAMINATED = {43, 55, 69, 80, 91, 110, 123}
print("Clearing contaminated rows...")
for r in sorted(CONTAMINATED):
    old = ws.cell(r, 2).value
    ws.cell(r, 2).value = None
    print(f"  Row {r}: cleared — was: {repr(str(old)[:80])}")

# ── 2. Fix Row 61 wording ──────────────────────────────────────────
print("\nUpdating Row 61...")
old61 = ws.cell(61, 2).value
ws.cell(61, 2).value = (
    "Used language appropriate to the customer's level "
    "and confirmed understanding throughout"
)
print(f"  was: {repr(old61)}")
print(f"  now: {repr(ws.cell(61, 2).value)}")

wb.save(FILE)
print("\nDone.")
