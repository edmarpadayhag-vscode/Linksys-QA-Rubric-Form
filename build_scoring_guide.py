
"""
Creates a comprehensive "Scoring Guide" sheet in QA Scoring Form Template.xlsx.
Each behavioral indicator gets a 4-row block: Met / Partial / Missed / N.A.
with explanation + example scenario for each rating.
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
from openpyxl.utils import get_column_letter

FILE  = r"d:\Visual Studio Code\QA Rubric\QA Scoring Form Template.xlsx"

# ── Color palette ─────────────────────────────────────────────────
C_TITLE_BG   = "1F3864"   # dark navy  – sheet title
C_TITLE_FG   = "FFFFFF"
C_SEC_BG     = "2E75B6"   # medium blue – section headers
C_SEC_FG     = "FFFFFF"
C_IND_BG     = "DEEAF1"   # pale blue   – indicator label
C_MET_BG     = "E2EFDA"   # pale green  – Met
C_PAR_BG     = "FFF2CC"   # pale yellow – Partial
C_MIS_BG     = "FCE4D6"   # pale salmon – Missed
C_NA_BG      = "F2F2F2"   # light grey  – N/A
C_HDR_BG     = "4472C4"   # bright blue – column headers

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic)

def align(wrap=True, v="top", h="left"):
    return Alignment(wrap_text=wrap, vertical=v, horizontal=h)

thin = Side(style="thin", color="BFBFBF")
def border():
    return Border(left=thin, right=thin, top=thin, bottom=thin)

# ── Guide content ─────────────────────────────────────────────────
# Each section: ('Section label', [ (indicator_text, met_exp, met_ex, par_exp, par_ex, mis_exp, mis_ex, na_text), ... ])
GUIDE = [
  ("1 · RESOLUTION / PROGRESS TOWARD RESOLUTION", [

    ("Resolved the issue or achieved the best possible outcome\n(RMA, valid escalation, limitation explained, education)",
     "Issue fully resolved and confirmed with the customer. Alternatively, the strongest available outcome was properly executed: RMA correctly initiated, valid escalation submitted with complete documentation, product limitation accurately explained, or appropriate customer education provided.",
     "Customer reported no internet on Velop node. Agent guided 5-press reset and re-pairing. Node showed solid white. Customer confirmed internet restored and agent verified before closing. → Met.",
     "An outcome was pursued but not fully executed or confirmed — RMA initiated but missing a required field, escalation submitted with sparse notes, or issue appeared resolved but the customer did not confirm.",
     "Agent guided a factory reset. Node appeared to re-pair. Agent closed the ticket without asking the customer to test internet. Customer did not confirm. → Partial.",
     "Case closed without any meaningful outcome. Generic KB link sent without investigation. Escalation used with no prior troubleshooting. Out-of-warranty product dismissed with no attempt to help.",
     "Customer called about no internet. Agent immediately sent a KB article without asking any questions, then marked the ticket Resolved. → Missed.",
     "Not applicable. Every interaction must produce some form of documented outcome or attempt."),

    ("Completed reasonable troubleshooting before reaching a conclusion",
     "Agent worked through relevant diagnostic steps — appropriate to the product and symptom — before arriving at any conclusion. Steps were logical, sequenced, and evidence-based.",
     "Customer reported intermittent drops. Agent checked for IP conflicts, verified firmware version, tested wired vs. wireless, confirmed drops were wireless-only, identified channel congestion and changed channels. → Met.",
     "Some troubleshooting was done but key steps were skipped. Conclusion reached without fully validating all likely causes.",
     "Customer reported no internet. Agent rebooted the router without first verifying the modem had a working WAN connection. The modem was offline. → Partial.",
     "No meaningful troubleshooting performed. Agent jumped to a conclusion, blamed ISP without evidence, or escalated after a single failed step.",
     "Customer reported slow speeds. Agent immediately told the customer to contact their ISP without running any diagnostics. → Missed.",
     "Not applicable. Troubleshooting is always possible or its absence must be documented."),

    ("Resolution path was appropriate for the situation and warranty status",
     "Agent correctly determined warranty status and recommended the right resolution path for the situation: in-warranty hardware failure → RMA; out-of-warranty → self-help or paid support; software or config issue → troubleshooting; known limitation → education.",
     "Customer reported a hardware failure 8 months after purchase. Agent verified in-warranty status, confirmed the fault through troubleshooting, and correctly initiated an RMA. → Met.",
     "Resolution path was mostly right but one element was misaligned — offered paid support when product was in warranty, or denied help to an in-warranty customer citing an incorrect out-of-warranty determination.",
     "Customer had an in-warranty device with a confirmed hardware fault. Agent correctly identified the issue but initially offered paid remote support before being corrected. → Partial.",
     "Completely wrong path taken — free replacement for a clearly out-of-warranty product; denied assistance to in-warranty customer; escalated a simple config issue without troubleshooting.",
     "Agent told the customer their 3-year-old router was still under warranty and initiated an RMA that would later be rejected. → Missed.",
     "N/A only when warranty status was genuinely impossible to determine despite a reasonable lookup, and agent documented that limitation."),

    ("Customer was moved meaningfully closer to a solution",
     "The interaction produced real, documented progress — the issue was resolved, a specific actionable next step was executed, or the customer's situation clearly advanced (RMA in process, escalation queued with full context, specific fix confirmed).",
     "Issue was not fully resolved but agent identified the root cause, submitted a complete escalation to the right team, and customer understood the next step. → Met.",
     "Some progress was made but the outcome was unclear or the next step was vague. Customer was helped but left without a clear picture of where things stand.",
     "Agent troubleshot and suggested the customer 'monitor and see if it happens again' with no timeframe, no trigger condition, and no follow-up plan. → Partial.",
     "Customer ends the interaction in the same state as when they called — no resolution, no clear next step, no advancement toward a fix.",
     "After a 30-minute call, agent concluded 'it might be your ISP' with no evidence and no follow-up plan, then closed the ticket. → Missed.",
     "Not applicable."),
  ]),

  ("2 · TECHNICAL ACCURACY", [

    ("Correctly identified the symptoms and asked relevant diagnostic questions",
     "Agent identified the specific symptom — not just 'internet issue' — and asked targeted follow-up questions to narrow scope: which devices affected, wired vs. wireless, when it started, any recent changes.",
     "Customer said 'internet is down.' Agent asked: all devices or one? Wired or wireless? When did it start? Anything change recently? Identified the issue was wireless-only on one device, not a full outage. → Met.",
     "Symptoms partially identified but agent asked generic questions without building on the answers; or missed a key clarifying question that would have changed the diagnostic path.",
     "Customer reported 'internet issues.' Agent asked for the router model and ran a standard reset without determining whether the problem was connectivity, speed, or intermittent. → Partial.",
     "Agent proceeded with generic troubleshooting without understanding the actual symptom; or misidentified the symptom and pursued the wrong path entirely.",
     "Customer reported slow speeds. Agent assumed it was a full outage and guided a factory reset. The actual issue was a speed tier mismatch with the ISP. → Missed.",
     "Customer immediately requested escalation and declined any diagnostic interaction."),

    ("Identified the root cause (or likely cause) when possible",
     "Based on evidence gathered, agent correctly identified the root cause or articulated the most likely cause with supporting observations — not a guess, but a conclusion drawn from test results and symptoms.",
     "After finding wired connection fine but wireless dropping, and noting drops coincided with device movement, agent concluded: likely signal interference or channel congestion on the 2.4 GHz band — supported by evidence. → Met.",
     "Agent identified a possible cause but could not narrow it down further, or stated a cause without supporting evidence. Cause was speculative.",
     "Agent said 'it's probably the firmware' without checking the firmware version or running any test to support that conclusion. → Partial.",
     "Agent failed to identify a cause or stated an unsupported cause — ISP blame, customer-side blame — without running any diagnostic to justify it.",
     "After a single failed ping test, agent concluded 'it must be your ISP' without checking modem status, WAN connection, or any device-side diagnostics. → Missed.",
     "Issue resolved without a definitive cause identifiable (e.g., self-clearing intermittent fault); or customer declined further investigation after an initial fix attempt."),

    ("Followed a logical troubleshooting process",
     "Steps built on each other in a structured sequence — simple checks before complex changes, verification before action, each step building on the previous result. No skipping foundational checks.",
     "Agent verified modem online → confirmed WAN port connection → opened router admin page → identified duplicate IP assignment → fixed with static IP. Clear progression from ISP side inward. → Met.",
     "Generally logical but steps were slightly out of order or a key verification step was skipped.",
     "Agent guided a full factory reset without first verifying the modem had a working internet connection. The actual problem was the modem, which L2 identified immediately. → Partial.",
     "Troubleshooting was random, reversed (complex before simple), or repeated unnecessarily with no response to the outcomes.",
     "Agent had the customer reset the router three times, then recommended a firmware update, then another factory reset — no logic, no response to results at each step. → Missed.",
     "No troubleshooting was applicable — product end-of-life with no available fix, or customer declined."),

    ("Available tools were used efficiently and appropriately",
     "Agent used relevant tools at the right point in the process — remote session when hands-on was needed, admin dashboard for logs, modem test for WAN verification, KB for procedure confirmation — and interpreted results correctly.",
     "Agent remotely accessed the router admin panel, reviewed the DHCP lease table and event logs, identified a firmware error code, cross-referenced the KB, and pinpointed the fix without requiring the customer to navigate menus. → Met.",
     "Agent used some tools but missed relevant ones; or used a tool at the wrong moment; or initiated a tool but could not correctly interpret the results.",
     "Agent ran a speed test (appropriate) but did not check event logs or the admin panel, which would have shown the firmware error causing the speed issue. → Partial.",
     "Agent did not use any available tools despite the situation requiring them; or used a tool that was irrelevant to the reported symptom.",
     "Customer reported frequent node drops. Agent only asked the customer to visually describe the LEDs rather than using the remote session or admin panel already available. → Missed.",
     "No tools were available or applicable — offline issue prevented tool access; or customer declined remote assistance."),

    ("Gave technically accurate information; no unsupported conclusions",
     "All technical guidance, product specifications, and procedural instructions were factually correct and aligned with KB documentation. Agent validated assumptions before presenting conclusions.",
     "Agent correctly described the Velop factory reset as 10 seconds, confirmed the LED sequence per KB, and matched the admin password to the product label. All verified before advising. → Met.",
     "Most information was accurate but one or two minor inaccuracies were present — slightly wrong reset duration, incorrect LED color description — unlikely to cause serious harm but reflecting a knowledge gap.",
     "Agent said the reset required 'about 10 to 15 seconds' (KB specifies 10). Customer held the button too long and triggered a factory reset when a soft reset was intended. → Partial.",
     "Agent provided materially incorrect information — wrong default password, non-existent LED state, unsupported pairing method, or made claims without evidence (ISP blame, customer-side blame).",
     "Agent told the customer the default admin password was 'admin admin.' Customer could not access the admin panel. The correct credential was printed on the device label. → Missed.",
     "Not applicable. Technical accuracy is always required."),
  ]),

  ("3 · COMMUNICATION & CALL CONTROL", [

    ("Expectations were set clearly from the start",
     "Agent acknowledged the issue, outlined what the interaction would involve, and gave a realistic estimate of time and likely outcome at the opening of the call. Customer knew what to expect before troubleshooting began.",
     "'I can see you're having connectivity issues. I'm going to walk you through a few diagnostic steps — should take about 10–15 minutes. Our goal is to get you back online today or confirm the next step.' → Met.",
     "Agent acknowledged the issue but gave no indication of what to expect in terms of process, time, or possible outcome.",
     "Agent said 'I'll help you with that' and immediately started troubleshooting without framing the interaction. → Partial.",
     "Agent went directly into troubleshooting with no introduction, no framing, and no expectation-setting of any kind.",
     "Agent's first statement was 'Have you tried restarting your router?' with no acknowledgment of the issue or explanation of what was about to happen. → Missed.",
     "Not applicable. Expectations can always be set, even briefly."),

    ("Agent maintained call control and troubleshooting flow",
     "Agent kept the conversation on track, redirected tangents professionally, paced the troubleshooting steps deliberately, and ensured the diagnostic path remained purposeful throughout.",
     "When the customer started discussing a separate device issue, agent said: 'That's worth looking at — let's note that and address it right after we sort the connectivity issue. Can you try the reset now?' → Met.",
     "Agent generally maintained control but lost direction at one or two points — allowed a long tangent, let the customer perform steps out of order, or lost momentum mid-call.",
     "Agent allowed the customer to spend 10 minutes describing their internet history before redirecting. The call ran significantly long as a result. → Partial.",
     "Customer dominated the interaction. Agent was reactive throughout, allowed the customer to dictate the troubleshooting path, or the interaction lacked any meaningful structure.",
     "Customer insisted 'just replace it.' Agent agreed to submit an RMA without completing any troubleshooting, despite the issue being a misconfigured setting fixable in minutes. → Missed.",
     "Not applicable."),

    ("Use appropriate language, as needed, and confirmed customer understanding",
     "Agent adapted terminology to the customer's demonstrated level — explained jargon when used, avoided unnecessary technical terms with non-technical customers, and verified comprehension at key transition points.",
     "For an elderly customer, agent said 'Find the small button on the back of the router — it may say Reset next to it. Let me know when you find it.' Rather than 'Locate the factory reset pinhole adjacent to the WAN port.' → Met.",
     "Language was mostly appropriate but agent used unexplained jargon on one or two occasions, or confirmed understanding only at the end rather than throughout.",
     "Agent used the term 'DHCP lease' without explanation. Customer nodded along but was clearly confused. Agent did not notice and moved on. → Partial.",
     "Agent used technical language throughout with no adaptation. Customer was visibly confused at multiple points and agent never addressed the comprehension gaps.",
     "Agent walked a non-technical customer through MAC filtering and DNS settings without checking whether the customer understood what these terms meant. Customer followed instructions incorrectly. → Missed.",
     "Not applicable."),

    ("Actively listened and showed comprehension throughout the interaction/s",
     "Agent demonstrated active listening — accurately reflected back what the customer said, built each question on the previous answer, and never asked for information the customer had already provided.",
     "Customer mentioned the issue started after a power outage. Later, agent said: 'Since you mentioned the power outage, let's check whether the router's settings were affected — that can happen when power is cut abruptly.' → Met.",
     "Agent generally listened but repeated a question already answered, or moved forward without fully incorporating a key detail the customer had shared.",
     "Customer mentioned they had already rebooted the router three times. Agent's next instruction was to reboot the router — without acknowledging this. → Partial.",
     "Agent repeatedly asked for information already provided, talked over the customer, or failed to incorporate what the customer said into the troubleshooting approach.",
     "Customer provided the router model, purchase date, and symptom upfront. Agent then asked for the router model, purchase date, and symptom again as if the customer had said nothing. → Missed.",
     "Not applicable."),

    ("Managed time well, with clear transitions between steps",
     "Interaction was efficient. Holds were managed with advance notice and updates. Agent transitioned clearly between troubleshooting phases using narration to keep the customer oriented.",
     "'We've confirmed the modem is working — good. Now let's look at the router. I'm going to put you on a brief hold to pull your device history — back in about 2 minutes.' → Met.",
     "Time management was generally adequate but there were noticeable silences, abrupt transitions between steps, or holds placed without warning.",
     "Agent placed the customer on hold without warning three times, returning each time to ask a new question rather than coming back with an answer. → Partial.",
     "Interaction ran well beyond what was necessary. Long unexplained silences. Customer left waiting with no update. Steps repeated unnecessarily.",
     "Agent left the customer on silent hold for 8 minutes with no update, returned to ask questions already answered, and the call ended unresolved after 45 minutes. → Missed.",
     "Not applicable."),

    ("Handled difficult moments when needed to maintain call control",
     "When the customer became frustrated, escalatory, or emotional, agent acknowledged the feeling empathetically, maintained composure, de-escalated professionally, and returned the interaction to productive troubleshooting.",
     "When customer said 'I've called three times and no one can fix this!' agent responded: 'I completely understand — that's incredibly frustrating, and I'm sorry. I'm going to review everything from your prior contacts and give this my full attention right now.' → Met.",
     "Agent attempted to handle the moment but de-escalation was mechanical or brief. Customer remained frustrated but interaction continued. Professionalism maintained.",
     "Customer was angry. Agent said 'I understand your frustration' and immediately moved to the next step without genuine acknowledgment. Customer stayed agitated. → Partial.",
     "Agent ignored frustration, matched negative energy, became defensive, or allowed the difficult moment to derail the interaction entirely.",
     "Customer raised their voice. Agent responded: 'I'm just trying to help you, there's no need to be rude.' Customer escalated further and requested a supervisor. → Missed.",
     "No difficult moments occurred — customer was cooperative throughout the interaction."),

    ("Adjusted communication for the customer's accessibility needs\n(hearing, vision, language barrier, age, cognitive)",
     "Agent identified or proactively accommodated an accessibility need and measurably adjusted delivery — slower speech, simpler vocabulary, avoiding visual-reference instructions, using plain step-by-step guidance suited to the customer's specific limitation.",
     "Customer mentioned they were hard of hearing. Agent slowed their pace, spoke clearly, emailed all instructions as numbered steps, and confirmed each step was understood before moving on. → Met.",
     "Agent partially accommodated — slowed down initially but reverted to standard pace; or addressed one aspect of the need but overlooked another.",
     "Agent noticed a language barrier and simplified vocabulary but still used terms like 'firmware' and 'SSID' without explanation. → Partial.",
     "Agent was aware of or should have been aware of an accessibility need and made no accommodation. Instructions were inaccessible or impractical for the customer.",
     "Customer mentioned they could not see the back of the router clearly. Agent continued giving visual-reference instructions ('the button is next to the yellow port') without offering an alternative. → Missed.",
     "No accessibility accommodation was required — customer had no communication barrier."),
  ]),

  ("4 · CUSTOMER OWNERSHIP", [

    ("Took accountability for the customer's outcome and avoided unnecessary transfers",
     "Agent owned the case from contact to close, made reasonable efforts to resolve or advance it personally, and only transferred when there was a clear, documented legitimate reason — wrong queue, specialized team required after L1 work was completed.",
     "Agent inherited a complex multi-touch case, reviewed all prior notes, re-diagnosed without making the customer repeat everything, and escalated only after completing thorough L1 troubleshooting with full documentation. → Met.",
     "Agent generally owned the case but transferred at a point where further effort was reasonable and expected; or the transfer was made with some context but the handoff documentation was incomplete.",
     "Agent attempted troubleshooting for 5 minutes and transferred to L2 citing complexity — the issue was a basic configuration problem clearly within L1 scope. → Partial.",
     "Agent transferred immediately or after minimal effort to avoid a difficult issue. Transfer happened without documentation or context for the receiving agent.",
     "Agent saw the case involved a mesh system and immediately routed to L2 without troubleshooting, noting only 'customer has Velop issue' in the ticket. → Missed.",
     "Not applicable. Ownership is always expected."),

    ("Commitments made were followed through, and a clear path forward was created for the customer.",
     "Every commitment made during the interaction was executed within the call or documented with a specific owner and timeline. Customer left with a defined, actionable next step.",
     "Agent committed to submitting the escalation before end of day, told the customer what L2 would do and when to expect contact, and updated the ticket accordingly before closing. → Met.",
     "Path forward was defined but vague — no specifics on who, when, or what. Or a commitment was made verbally but not documented in the case.",
     "Agent told the customer 'someone will follow up with you' but did not specify the team, timeline, or the nature of the follow-up. → Partial.",
     "Agent made a commitment that was not kept and not documented. Customer ends the interaction with no defined next step.",
     "Agent promised to call back within 24 hours. No callback was made and no note was left in the ticket. Customer called back the next day. → Missed.",
     "Not applicable. A path forward is always relevant."),

    ("Clearly set customer expectations for next steps and realistic timelines",
     "Agent told the customer specifically what would happen after the interaction — who would contact them, when, what to do while waiting, and any conditions or contingencies. Timelines given were realistic and achievable.",
     "'Your case is with our Level 2 team. They will reach out within 24–48 business hours at the number on file. Please don't reset the device again in the meantime — that will help them diagnose it when they call.' → Met.",
     "Some expectation-setting occurred but it was incomplete — timeline provided without a team name, or team named without a timeframe, or the customer's own role in the next step was unclear.",
     "Agent said 'L2 will reach out to you' with no indication of when, how, or what the customer should prepare. → Partial.",
     "Customer received no information about what happens next. Or was given a timeline that was clearly unrealistic and would not be met.",
     "Agent told the customer they would hear back 'very soon' for an escalation that typically takes 3–5 business days. → Missed.",
     "Not applicable. Expectations should always be set."),

    ("Completed required disconnect follow-up and met callback commitments within the agreed window",
     "If the call disconnected, agent made a documented callback attempt within the expected window. If a callback was promised at a specific time, it was made on time — or proactively renegotiated before the window lapsed.",
     "Call dropped mid-troubleshooting. Agent immediately called back, reached the customer, resumed from where they left off, and noted the disconnect and callback in the ticket. → Met.",
     "Follow-up was attempted but documentation was incomplete; or callback was made but slightly outside the agreed window without proactive communication to the customer.",
     "Agent called back after a disconnect but did not update the ticket, leaving no record of the follow-up for the next agent. → Partial.",
     "Call disconnected and no follow-up was made or documented. Callback was promised and not delivered.",
     "Call dropped at a critical point. Agent did not call back. Customer called in the following day and had to start over with a new agent who had no record of the prior session. → Missed.",
     "No disconnect occurred during the interaction and no callback was committed."),
  ]),

  ("5 · ESCALATION JUDGMENT", [

    ("Completed reasonable troubleshooting before escalating",
     "Agent completed all relevant L1 troubleshooting steps applicable to the issue — sufficient to establish that L2 intervention was genuinely needed. Steps documented before escalating.",
     "Customer's mesh node was stuck on red after multiple reset attempts across two contacts. Agent completed all standard recovery steps, documented results, and escalated with full history. → Met.",
     "Some troubleshooting was completed but one or two standard L1 steps were skipped that could have resolved the issue or provided clearer diagnostic data for L2.",
     "Agent ran a factory reset but did not check modem connectivity first. Escalated citing 'hardware failure' — L2 immediately identified the modem as the actual culprit. → Partial.",
     "Agent escalated immediately or after minimal effort. L2 received a case where basic troubleshooting had not been attempted.",
     "Customer reported no internet. Agent ran one restart, found the issue unresolved, and immediately escalated to L2 with no modem check and no other diagnostic steps. → Missed.",
     "No escalation occurred during this interaction."),

    ("Identified a valid reason to escalate",
     "Escalation criterion was clearly met — technical complexity beyond L1 scope, confirmed hardware failure, repeat unresolved issue with documented history, customer complaint or legal-risk trigger, or issue requiring L2 tools or authority.",
     "After two contacts and complete L1 troubleshooting, the node consistently failed to re-pair. Agent identified likely hardware defect, documented the repeated failure, and escalated with evidence. → Met.",
     "Reason for escalation was present but borderline — issue may have been resolvable at L1 with more effort; or the escalation trigger was identified but not clearly documented.",
     "Agent escalated citing 'customer is frustrated' — a secondary factor, but no technical criterion was articulated and the issue was within L1 scope. → Partial.",
     "No valid escalation criterion was met. Agent escalated to avoid a difficult call, or escalated a routine issue clearly within L1 scope without any supporting rationale.",
     "Customer asked for an escalation. Agent complied immediately without completing any troubleshooting or documenting a technical reason. → Missed.",
     "No escalation occurred during this interaction."),

    ("Provided complete escalation details using the correct escalation path",
     "Escalation contained: specific symptom description, all troubleshooting steps and their results, root cause hypothesis, reason for escalation, customer contact details, case history reference, and was submitted to the correct team or queue.",
     "Escalation note: 'Customer reports node stuck on red LED after 3 reset attempts across 2 contacts. Modem confirmed online. Factory reset performed — node pairs briefly then drops. Likely hardware defect. All steps in prior case #12345. Submitted to Tier 2 Hardware.' → Met.",
     "Escalation submitted to the right team but missing some useful detail — no test results, no case history, no contact info. L2 can action it but may need to re-contact the customer for context.",
     "Escalation note: 'Customer has connectivity issue. Troubleshooting performed. Escalating.' Missing results, steps, and contact. L2 had to call the customer to understand the issue. → Partial.",
     "Critical information missing — no troubleshooting details, no customer contact, no symptom description. Or escalated to the wrong team or queue entirely.",
     "Escalation note: 'Escalating per customer request.' No technical reason, no troubleshooting summary, no case history. L2 had to start from scratch. → Missed.",
     "No escalation occurred during this interaction."),

    ("Clearly explained why escalation was necessary and what would happen next.",
     "Agent explained in plain language why the case needed to go to L2 — what the issue required that L1 couldn't provide — what the escalation process looked like, and what the customer should realistically expect.",
     "'The tests suggest a hardware issue with the node that our Level 2 specialists are better equipped to assess. They'll reach out within 24–48 hours. If confirmed defective, they'll start the replacement process.' → Met.",
     "Agent told the customer about the escalation but didn't explain why, or gave a vague timeline without specifics about what L2 would actually do.",
     "Agent said 'I'm escalating this to our next level of support — they'll be in touch.' No explanation of why, no L2 action described, no timeline given. → Partial.",
     "Agent submitted the escalation without telling the customer, or told the customer the case was resolved when it was actually being escalated.",
     "Agent escalated the case, marked the ticket Resolved, and ended the call. Customer was confused when L2 called two days later. → Missed.",
     "No escalation occurred during this interaction."),
  ]),

  ("6 · CUSTOMER EXPERIENCE", [

    ("Demonstrated empathy including for repeat-contact fatigue",
     "Agent acknowledged the customer's frustration or history in a sincere, specific way. For repeat contacts, explicitly recognized the customer's repeated effort and validated their experience — not a scripted phrase.",
     "'I can see this is your third call about this issue — that must be incredibly frustrating. I want to make sure we get to the bottom of this today. I've reviewed your prior contacts so you won't need to start over.' → Met.",
     "Empathy was expressed but generic and scripted-sounding, without acknowledging the specific situation or the customer's contact history.",
     "Agent said 'I understand your frustration' and immediately moved to troubleshooting with no further acknowledgment of the customer's experience. → Partial.",
     "Agent was entirely transactional. No acknowledgment of frustration, no recognition of repeat contact history. Customer felt processed, not heard.",
     "Customer's fourth call about the same unresolved issue. Agent's first statement was 'Can I have your model number?' with no acknowledgment of the history. → Missed.",
     "Customer expressed no frustration and had no prior contact history requiring acknowledgment."),

    ("Stayed professional throughout and maintained patience, even under pressure",
     "Agent remained composed, courteous, and helpful throughout — even when the customer was frustrated, demanding, or uncooperative. Did not take provocations personally or let them alter tone.",
     "Customer was rude and demanding across a 45-minute call. Agent remained calm, professional, and solution-focused at every point without becoming defensive. → Met.",
     "Agent was mostly professional but had a brief moment of visible impatience or a slightly defensive response — noticeable but not significantly damaging to the interaction.",
     "After a long, difficult call, agent said 'I've been very patient with you today' in a tone that implied criticism. Customer filed a complaint. → Partial.",
     "Agent became visibly impatient, sarcastic, dismissive, or defensive in a way that materially damaged the interaction or made the customer feel disrespected.",
     "Customer raised their voice. Agent responded: 'I'm just trying to help you, there's no need to be rude.' Customer escalated further. → Missed.",
     "Not applicable. Professionalism is always required."),

    ("Agent matched the customer's tone, pace, level, and energy",
     "Agent adapted naturally to the customer's communication style — technical and precise with a knowledgeable customer, warm and patient with an anxious first-time user, efficient and direct with an experienced caller who wanted a quick resolution.",
     "Customer identified themselves as a network engineer. Agent shifted to technical terminology, skipped basic explanations, and engaged as a peer. Customer appreciated the efficient, direct exchange. → Met.",
     "Agent mostly matched the customer but one dimension was off — too formal with a casual customer, too technical with someone clearly out of their depth, or too slow with an experienced user who wanted speed.",
     "Customer was casual and relaxed. Agent maintained a stiff, formal tone throughout, creating a slight but noticeable disconnect even though the technical help was sound. → Partial.",
     "Agent maintained a one-size-fits-all tone throughout regardless of the customer's cues. Interaction felt scripted and impersonal.",
     "Elderly customer was nervous and uncertain. Agent delivered fast, jargon-heavy instructions as if speaking to a technical audience. Customer became more confused. → Missed.",
     "Not applicable. Tone matching is always relevant."),

    ("Minimized customer effort",
     "Agent proactively reduced what the customer needed to do — did not ask for information already known, handled actions agent-side where possible, confirmed information before asking the customer to perform steps.",
     "Agent had the device model and serial from the account before the call and only asked the customer to confirm the current symptom — saving several minutes and avoiding repetitive questions. → Met.",
     "Agent made some effort to reduce customer effort but had areas where the customer was asked to repeat known information or perform steps that could have been handled more efficiently.",
     "Agent asked the customer for the device model that was already visible in the ticket from the prior interaction. → Partial.",
     "Customer had to repeat information multiple times, was sent through unnecessary steps, or experienced avoidable friction because the agent did not use available information.",
     "Customer on their third contact was asked by the agent to re-explain the issue and repeat all troubleshooting already performed in the prior two contacts. → Missed.",
     "Not applicable."),

    ("Maintained customer understanding and engagement",
     "Throughout the interaction, the customer remained engaged, understood what was happening and why, and knew what they were supposed to do at each step. Agent verified comprehension and addressed confusion immediately.",
     "Before each step, agent explained what they were looking for and why. After each step, confirmed the result. Customer was an active, informed participant. → Met.",
     "Customer was generally engaged but had moments of confusion that were only partially addressed; or the customer's engagement dropped and the agent did not notice or respond.",
     "Customer said 'sorry, I'm a bit lost.' Agent repeated the last instruction without simplifying or reframing. Customer continued but remained uncertain. → Partial.",
     "Customer was disengaged or lost at multiple points. Agent continued troubleshooting without addressing the comprehension gaps. Customer was sidelined in their own support call.",
     "Agent performed all steps remotely without explaining what was being done. Customer had no idea what was happening and did not feel involved. → Missed.",
     "Not applicable."),
  ]),

  ("7 · DOCUMENTATION & NOTES", [

    ("Recorded customer issue summary",
     "Case notes clearly describe what the customer reported — specific symptom, affected device(s), scope of impact, when it started, and what the customer had already tried before calling.",
     "'Customer reports wireless drops on all 3 Velop nodes. Issue started 3 days ago after an overnight firmware update. Customer has rebooted all nodes. Wired connection to main node is unaffected.' → Met.",
     "Issue summary present but vague — 'customer called about connectivity issue' without specific symptom, scope, or prior steps. Next agent cannot tell what was actually wrong.",
     "Case note: 'Customer experiencing internet issues with their router. Troubleshooting performed.' → Partial.",
     "No issue summary recorded at all. Next agent has no starting context.",
     "Case closed with only 'Resolved' in the notes — no description of what the customer reported or what was wrong. → Missed.",
     "Not applicable."),

    ("Documented troubleshooting steps",
     "Notes clearly list troubleshooting steps performed during the interaction in a logical sequence — what was done, in what order, with what result at each step.",
     "'1. Verified modem online — WAN light confirmed on. 2. Checked DHCP lease — IP assigned correctly. 3. Ran wireless diagnostic — signal weak on 2.4 GHz. 4. Changed Wi-Fi channel to channel 6. 5. Customer tested — connectivity restored.' → Met.",
     "Some steps noted but the list is incomplete — key steps missing, or outcomes of individual steps not recorded.",
     "'Reset router and reconfigured Wi-Fi settings.' No detail on which settings, what the reset revealed, or the outcome of each action. → Partial.",
     "No troubleshooting steps documented despite a full troubleshooting session. Next agent cannot tell what was tried.",
     "After a 45-minute call: case note reads 'Troubleshooting performed. Escalated.' Nothing else. → Missed.",
     "No troubleshooting was performed — call was informational only, or customer declined."),

    ("Noted key findings",
     "Notes capture what was discovered — test results, error messages observed, device behavior, what worked and what did not. Findings are specific and actionable for the next agent.",
     "'Admin panel shows firmware version 3.8.1 — latest is 4.1.2. Firmware update failed twice. Event log shows error code E404: firmware download timeout. Likely firmware corruption or server-side download issue.' → Met.",
     "Some findings noted but key observations are missing — outcome of a test not recorded, error message observed but not documented, inconsistency noticed but not noted.",
     "Notes mention a factory reset was performed but don't record what happened afterward or whether the issue persisted. → Partial.",
     "No findings documented at all. Next agent cannot tell what was learned from the interaction.",
     "Notes record only actions. No results, no observations, no test outcomes. Next agent has to start from scratch. → Missed.",
     "No diagnostic findings to record — quick informational inquiry with no testing involved."),

    ("Captured agent reasoning — escalation justification, offline recommendations provided by colleagues and other support POCs especially if no private notes applicable to step/action",
     "Notes explain key decisions — why a specific troubleshooting path was chosen, why the case was escalated (what criterion was met), and any colleague or SME consultations that shaped the approach. If no private notes are available, colleague input is captured in the case notes.",
     "'Consulted with Team Lead before escalating. TL confirmed this meets the hardware defect threshold after three failed factory resets. Escalated to Tier 2 Hardware with full history. TL input documented here as no private note was available.' → Met.",
     "Some reasoning captured but key decisions lack explanation — escalation submitted with no documented rationale, or colleague input followed but not attributed in any note.",
     "Agent escalated and noted 'escalated per policy' but did not document which criterion was met, what troubleshooting led to that conclusion, or who they consulted. → Partial.",
     "Notes contain only actions — what was done — with no explanation of why. Next agent or QA reviewer cannot understand the decision-making.",
     "Agent escalated the case. Escalation note: 'Escalating to L2.' No reason, no troubleshooting summary, no colleague reference. → Missed.",
     "Straightforward interaction with no escalation, no colleague consultation, and no ambiguous decisions requiring explanation."),

    ("Documented next steps and follow-ups",
     "Notes specify what happens after the interaction — who is responsible for the next action, what that action is, and any relevant timeline or conditions. Callback details fully recorded where applicable.",
     "'Escalated to L2 Hardware (ticket #67890). Customer to be contacted within 24–48 hours at 555-0100. Customer advised not to factory reset again until L2 makes contact.' → Met.",
     "Next steps mentioned in general terms but lacking specifics — no owner, no timeline, no conditions.",
     "'Escalated. Customer to wait for follow-up.' No team named, no timeframe, no detail on what the follow-up entails. → Partial.",
     "No next steps documented. Case cannot be picked up by another agent without re-contacting the customer.",
     "Case closed after escalation with no note on what was submitted, which team received it, or when the customer should expect contact. → Missed.",
     "Issue fully resolved with no pending actions required."),

    ("Updated ticket status accurately",
     "Ticket status accurately reflects the true state of the case at close — Resolved only when the issue was confirmed resolved, Escalated when escalated, Callback when a return call is pending. Closure reason documented.",
     "Issue was unresolved at close. Agent marked the ticket 'Escalated,' noted the receiving team and escalation ticket number, and documented the current issue state before closing. → Met.",
     "Status is mostly accurate but the closure reason is missing or too brief; or the correct status was used but applied late.",
     "Ticket was marked 'Resolved' with a closure note that said only 'fixed' — no description of what was done, confirmed, or by whom. → Partial.",
     "Ticket marked Resolved when the issue was not resolved. Or marked Closed without a closure reason. Or left in a status that would mislead the next agent.",
     "Agent submitted an escalation but marked the ticket 'Resolved' because the call ended. L2 received the case with a Resolved status and deprioritized it. → Missed.",
     "Not applicable. Ticket status must always reflect reality."),
  ]),
]

# ── Build the sheet ───────────────────────────────────────────────
wb = load_workbook(FILE)

# Remove existing Scoring Guide sheet if present
if "Scoring Guide" in wb.sheetnames:
    del wb["Scoring Guide"]

ws = wb.create_sheet("Scoring Guide")

# Column widths
col_widths = {"A": 30, "B": 44, "C": 11, "D": 62, "E": 62}
for col, w in col_widths.items():
    ws.column_dimensions[col].width = w

# ── Title row ────────────────────────────────────────────────────
ws.merge_cells("A1:E1")
c = ws["A1"]
c.value = "QA SCORING GUIDE — Behavioral Indicator Rating Definitions & Scenarios"
c.font = font(bold=True, color=C_TITLE_FG, size=13)
c.fill = fill(C_TITLE_BG)
c.alignment = align(h="center", v="center")
ws.row_dimensions[1].height = 28

# ── Sub-title ────────────────────────────────────────────────────
ws.merge_cells("A2:E2")
c = ws["A2"]
c.value = ("Met = fully achieved  ·  Partial = achieved in part  ·  "
           "Missed = not achieved  ·  N/A = not applicable to this interaction")
c.font = font(italic=True, size=9, color="FFFFFF")
c.fill = fill("3F3F3F")
c.alignment = align(h="center", v="center")
ws.row_dimensions[2].height = 18

# ── Column headers ───────────────────────────────────────────────
headers = ["Category / Section", "Behavioral Indicator", "Rating",
           "What This Looks Like", "Example Scenario"]
for col_idx, hdr in enumerate(headers, 1):
    c = ws.cell(3, col_idx)
    c.value = hdr
    c.font = font(bold=True, color=C_TITLE_FG, size=10)
    c.fill = fill(C_HDR_BG)
    c.alignment = align(h="center", v="center")
    c.border = border()
ws.row_dimensions[3].height = 22

# ── Rating rows setup ────────────────────────────────────────────
RATINGS     = ["Met", "Partial", "Missed", "N/A"]
RATING_FILL = [C_MET_BG, C_PAR_BG, C_MIS_BG, C_NA_BG]
RATING_FONT_COLOR = ["1A5E20", "7B4F00", "7B1919", "3D3D3D"]

current_row = 4
section_count = 0

for section_label, indicators in GUIDE:
    section_count += 1
    num_indicators = len(indicators)
    block_rows = num_indicators * 4   # 4 rating rows per indicator

    # Section header — merged across all indicator rows for this section
    sec_start = current_row
    sec_end   = current_row + block_rows - 1
    ws.merge_cells(start_row=sec_start, start_column=1,
                   end_row=sec_end,   end_column=1)
    c = ws.cell(sec_start, 1)
    c.value = section_label
    c.font = font(bold=True, color=C_SEC_FG, size=10)
    c.fill = fill(C_SEC_BG)
    c.alignment = align(h="center", v="center")
    c.border = border()

    for ind_text, met_exp, met_ex, par_exp, par_ex, mis_exp, mis_ex, na_text in indicators:
        ind_start = current_row
        ind_end   = current_row + 3

        # Indicator label — merged across 4 rating rows in column B
        ws.merge_cells(start_row=ind_start, start_column=2,
                       end_row=ind_end,     end_column=2)
        c = ws.cell(ind_start, 2)
        c.value = ind_text
        c.font = font(bold=True, size=9.5)
        c.fill = fill(C_IND_BG)
        c.alignment = align(v="center")
        c.border = border()

        # The 4 rating rows
        explanations = [met_exp, par_exp, mis_exp, na_text]
        scenarios    = [met_ex,  par_ex,  mis_ex,  ""]

        for i, (rating, bg, fc, exp, scen) in enumerate(
                zip(RATINGS, RATING_FILL, RATING_FONT_COLOR, explanations, scenarios)):
            row = ind_start + i

            # Rating cell (col C)
            c = ws.cell(row, 3)
            c.value = rating
            c.font = font(bold=True, color=fc, size=9.5)
            c.fill = fill(bg)
            c.alignment = align(h="center", v="top")
            c.border = border()
            ws.row_dimensions[row].height = 72

            # Explanation cell (col D)
            c = ws.cell(row, 4)
            c.value = exp
            c.font = font(size=9.5)
            c.fill = fill(bg)
            c.alignment = align(v="top")
            c.border = border()

            # Scenario cell (col E)
            c = ws.cell(row, 5)
            c.value = scen if scen else ("" if rating == "N/A" else "")
            c.font = font(size=9.5, italic=True, color="1A1A5E" if scen else "888888")
            c.fill = fill(bg)
            c.alignment = align(v="top")
            c.border = border()

        current_row += 4

# Freeze panes below header rows
ws.freeze_panes = "A4"

print(f"Scoring Guide built — {current_row - 4} rows across {section_count} sections.")
print("Saving...")
wb.save(FILE)
print("Done.")
