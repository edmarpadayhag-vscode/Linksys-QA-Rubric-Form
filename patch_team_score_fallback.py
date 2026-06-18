
"""
Update Team Score formula (G56:G59) to use the L1/L2 split only when
both levels are present in a case. All-L1 or all-L2 cases fall back to
the original 0.5 × Avg Touch formula.

Logic:
  IF mixed (L1 count > 0 AND L2 count > 0):
      0.2 × L1_avg + 0.3 × L2_avg + 0.3 × (Outcome/5) + 0.2 × (Cont/5)
  ELSE (all one level, or no level tagged):
      0.5 × D  + 0.3 × (Outcome/5) + 0.2 × (Cont/5)
"""

from openpyxl import load_workbook

FILE  = r"d:\Visual Studio Code\QA Rubric\QA Scoring Form Template.xlsx"
SHEET = "Case Tracker"

wb = load_workbook(FILE)
ws = wb[SHEET]

def team_score_formula(r):
    has_l1 = f'COUNTIFS($B$5:$B$49,A{r},$Q$5:$Q$49,"Level 1")>0'
    has_l2 = f'COUNTIFS($B$5:$B$49,A{r},$Q$5:$Q$49,"Level 2")>0'
    l1_avg = f'IFERROR(AVERAGEIFS($L$5:$L$49,$B$5:$B$49,A{r},$Q$5:$Q$49,"Level 1"),0)'
    l2_avg = f'IFERROR(AVERAGEIFS($L$5:$L$49,$B$5:$B$49,A{r},$Q$5:$Q$49,"Level 2"),0)'
    return (
        f'=IF(OR(A{r}="",D{r}="",E{r}="",F{r}=""),"",IFERROR('
        f'IF(AND({has_l1},{has_l2}),'
        f'0.2*{l1_avg}+0.3*{l2_avg},'
        f'0.5*D{r})'
        f'+0.3*(E{r}/5)+0.2*(F{r}/5),""))'
    )

for r in (56, 57, 58, 59):
    formula = team_score_formula(r)
    ws.cell(r, 7).value = formula
    print(f"G{r}: {formula}")

wb.save(FILE)
print("\nDone.")
